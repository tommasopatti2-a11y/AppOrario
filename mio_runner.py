# ================================
# Setup ambiente (disabilitato in ambiente server)
# ================================
# NOTE: Le righe Colab vengono disabilitate per l'esecuzione nel backend FastAPI/Docker.
# Se necessario in ambiente notebook, riabilitare manualmente.
# !pip -q install pandas openpyxl
# from google.colab import drive
# drive.mount('/content/drive')

# ================================
# Import
# ================================
from pathlib import Path
import re
import pandas as pd
from collections import defaultdict, OrderedDict

# ================================
# Config & Percorsi
# ================================
DEBUG = False

BASE = Path("/content/drive/MyDrive/DADA_project/Genera_Stampe")
CENTRALE_XLSX         = Path("/content/drive/MyDrive/DADA_project/Genera_Stampe/Input/Centrale.xlsx")
SUCCURSALE_XLSX       = Path("/content/drive/MyDrive/DADA_project/Genera_Stampe/Input/Succursale.xlsx")
TABELLA_AULE_XLSX     = Path("/content/drive/MyDrive/DADA_project/Assegnazione_Aule/Globale/Tabella_Aule.xlsx")
TABELLA_CLASSI_XLSX   = Path("/content/drive/MyDrive/DADA_project/Genera_Stampe/Input/Tabella_Classi.xlsx")
TABELLA_MATERIE_XLSX  = Path("/content/drive/MyDrive/DADA_project/Genera_Stampe/Input/Tabella_Materie.xlsx")
TABELLA_SOSTEGNO_XLSX = Path("/content/drive/MyDrive/DADA_project/Genera_Stampe/Input/Tabella_Sostegno.xlsx")



# Scrive in locale (filesystem Colab) per evitare rallentamenti/sync di Drive
TMP_DIR = Path("./tmp_outputs")
TMP_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR = TMP_DIR  # tutte le funzioni usano questa variabile


# ================================
# Costanti tema / orari
# ================================


# mappa ora -> orario di inizio (espandibile)
ORE_MAP = {1:"08:00", 2:"09:00", 3:"10:00", 4:"11:00", 5:"12:00", 6:"13:00", 7:"14:00"}

# ================================
# Palette & helper GLOBALI (Excel)
# ================================

# Blu principale
PRIMARY_BLUE_HEX  = "1565C0"                 # Excel/openpyxl (senza '#')

# Alias retro-compatibilità
BLUE_HEX  = PRIMARY_BLUE_HEX                 # per Excel

# Se in giro usi questi nomi:
COLOR_EMPHASIS_BLUE_HEX  = PRIMARY_BLUE_HEX

# Colori header/griglia (Excel)
COLOR_HEADER_HEX    = "D9E3F0"
COLOR_GRID_HEX      = "808080"
COLOR_ZEBRA_ALT_HEX = "EEF3FB"

def XLSX(hex_str: str) -> str:
    s = str(hex_str).strip().lstrip("#")
    if len(s) == 3:
        s = "".join(ch*2 for ch in s)
    return s.upper()


# ================================
# Helper COMUNI per tutto il file
# ================================
import re
from openpyxl.styles import Alignment, Font, PatternFill
from pandas import ExcelWriter

def is_aula_token(tok: str, known_aule) -> bool:
    t = (tok or "").strip().lower()
    if not t: return False
    if t in known_aule: return True
    return bool(re.match(r"^[A-Za-z]{0,4}\d{2,4}$", t)) or t.startswith("lab") or t.startswith("aula ")

def norm_class_token(tok: str) -> str:
    """Per la sola visualizzazione: rimuove ^ iniziali e * finali."""
    s = (tok or "").strip()
    s = s.lstrip("^")
    s = re.sub(r"\*+$", "", s)
    return s

def build_room_lookup(df_all=None, df_aule=None):
    """
    Ritorna:
      - known_aule: set case-insensitive di aule note
      - room2plesso: dict aula.lower() -> 'C'/'S'/None
    """
    known_aule = set()
    if df_all is not None and "aula" in df_all.columns:
        known_aule |= {str(a).strip().lower() for a in df_all["aula"].dropna() if str(a).strip()}
    if df_aule is not None and "Aula" in df_aule.columns:
        known_aule |= {str(a).strip().lower() for a in df_aule["Aula"].dropna() if str(a).strip()}
    room2plesso = {}
    if df_aule is not None and {"Aula","Plesso"} <= set(df_aule.columns):
        for _, r in df_aule.iterrows():
            a = tidy(r.get("Aula","")); p = tidy(r.get("Plesso","")).lower()
            if not a: continue
            tag = "C" if "centr" in p else ("S" if "succ" in p else None)
            room2plesso[a.lower()] = tag
    return known_aule, room2plesso

def excel_get_fills():
    """Colori coerenti ovunque con fallback sicuri."""
    hdr = XLSX(COLOR_HEADER_HEX)    if "COLOR_HEADER_HEX"    in globals() else "D9E3F0"
    zebra = XLSX(COLOR_ZEBRA_ALT_HEX) if "COLOR_ZEBRA_ALT_HEX" in globals() else "EEF4FF"
    fill_header = PatternFill("solid", fgColor=hdr)
    fill_title  = PatternFill("solid", fgColor=hdr)
    alt_fill    = PatternFill("solid", fgColor=zebra)
    return fill_header, fill_title, alt_fill


def excel_add_global_header(ws, ncols: int):
    """Scrive le righe 1–2 con HEADER_TEXT, le merge e le colora come header."""
    text = (HEADER_TEXT or "") if "HEADER_TEXT" in globals() else ""
    ws.append([text]); ws.append([""])
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=1, column=1)
    c.font = Font(bold=True, italic=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    fill_header, _, _ = excel_get_fills()
    for r in (1, 2):
        for cc in range(1, ncols+1):
            ws.cell(row=r, column=cc).fill = fill_header


def _dispo_tag_from_classes_cell(cls_cell: str) -> str | None:
    toks = [t.strip() for t in split_tokens(tidy(cls_cell))]
    if any(t == "d" for t in toks): return "S"
    if any(t == "D" for t in toks): return "C"
    return None

def _build_class_plesso_tag(df_classi):
    """
    Ritorna: dict {classe_norm -> 'C'/'S'} usando Tabella_Classi.
    Accetta sia colonna 'Edificio' sia 'Plesso'.
    """
    mp = {}
    if df_classi is None:
        return mp
    pl_col = "Edificio" if "Edificio" in df_classi.columns else ("Plesso" if "Plesso" in df_classi.columns else None)
    if pl_col is None or "Classe" not in df_classi.columns:
        return mp
    for _, r in df_classi.iterrows():
        cls = _norm_lookup_classe(tidy(r.get("Classe","")))
        pl  = tidy(r.get(pl_col,"")).lower()
        if not cls or not pl:
            continue
        if "centr" in pl:
            mp[cls] = "C"
        elif "succ" in pl:
            mp[cls] = "S"
    return mp


def _write_single_sheet_xlsx(xlsx_path: Path, sheet_name: str, data_matrix: list[list]):
    """Scrive una matrice (lista di liste) in un unico foglio Excel."""
    df = pd.DataFrame(data_matrix[1:], columns=data_matrix[0])
    with ExcelWriter(xlsx_path, engine="openpyxl") as ew:
        df.to_excel(ew, index=False, sheet_name=sheet_name)

def load_tabella_sostegno(tabella_sostegno_path, sheet_name=0, df_aule=None):
    """
    Legge Tabella_Sostegno con header a 2 righe:
      riga1: 'DOCENTE', 'Lunedì', '', '', 'Martedì', ...
      riga2: '', '1', '2', '', '1', '2', ...
    Restituisce (df_sostegno, giorni_order, ore_order) con colonne:
      ['plesso','docente','classe','aula','giorno','ora']
    Dove 'aula' è vuota (verrà riempita in integrate_sostegno_and_mark).
    """
    import pandas as pd, re
    from pathlib import Path

    # ---- helpers locali (non dipendono dal resto del file) ----
    DAY_ALIASES = {
        "lun":"Lunedì","lunedi":"Lunedì","lunedì":"Lunedì",
        "mar":"Martedì","martedi":"Martedì","martedì":"Martedì",
        "mer":"Mercoledì","mercoledi":"Mercoledì","mercoledì":"Mercoledì",
        "gio":"Giovedì","giovedi":"Giovedì","giovedì":"Giovedì",
        "ven":"Venerdì","venerdi":"Venerdì","venerdì":"Venerdì",
        "sab":"Sabato","sabato":"Sabato","dom":"Domenica","domenica":"Domenica",
    }
    def tidy(x):
        if pd.isna(x): return ""
        s = str(x).strip()
        return "" if s.lower()=="nan" else s
    def _norm_day(tok: str) -> str:
        t = (tok or "").strip().lower()
        return DAY_ALIASES.get(t, (tok or "").strip())
    def split_tokens(s: str):
        if not s: return []
        s = str(s).strip()
        for a,b in re.findall(r"\[(.*?)\]|\((.*?)\)", s):
            tok = (a or b or "").strip()
            if tok: s += f" | {tok}"
        return [p.strip() for p in re.split(r"\s*[\|/–—-]\s*", s) if p.strip()]
    def _is_aula_token(tok: str, known_aule) -> bool:
        t = (tok or "").strip().lower()
        if not t: return False
        if t in known_aule: return True
        return bool(re.match(r"^[A-Za-z]{0,4}\d{2,4}$", t)) or t.startswith("lab") or t.startswith("aula ")

    # ---- trova la riga header (prima colonna = DOCENTE) ----
    raw = pd.read_excel(tabella_sostegno_path, sheet_name=sheet_name, header=None, dtype=str, engine="openpyxl")
    header_idx = None
    for i in range(min(10, len(raw))):
        v = tidy(raw.iat[i,0]) if raw.shape[1] else ""
        if v.lower() == "docente":
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"{Path(tabella_sostegno_path).name}: non trovo la colonna 'Docente' nelle prime 10 righe.")

    # ---- leggi con header a due righe ----
    df = pd.read_excel(tabella_sostegno_path, sheet_name=sheet_name,
                       header=[header_idx, header_idx+1], dtype=str, engine="openpyxl")

    # aule note (per poter scartare eventuali token aula nelle celle classi)
    known_aule = set()
    if df_aule is not None and "Aula" in df_aule.columns:
        known_aule = {str(a).strip().lower() for a in df_aule["Aula"].dropna() if str(a).strip()}

    # ---- costruisci nomi colonne 'Giorno_Ora' propagando il giorno e numerando ore mancanti ----
    newcols = []
    prev_day = ""
    per_day_counter = {}
    for a, b in df.columns:
        A = tidy(a); B = tidy(b)
        # colonna Docente
        if A.lower() == "docente" or B.lower() == "docente":
            newcols.append("Docente")
            continue
        day = _norm_day(A) if A else ""
        # gestisci 'Unnamed: n' propagando il giorno precedente
        if (not day) or day.lower().startswith("unnamed"):
            day = prev_day
        if day:
            prev_day = day
        # ora: prendi numero nella 2a riga; altrimenti conta progressivo
        m = re.search(r"(\d+)", B)
        if m:
            hour = int(m.group(1))
        elif day:
            per_day_counter[day] = per_day_counter.get(day, 0) + 1
            hour = per_day_counter[day]
        else:
            # colonna inutile/di servizio
            newcols.append(f"{A or B}".strip() or "col")
            continue
        newcols.append(f"{day}_{hour}")

    df.columns = newcols

    # ---- estrai record long: (docente, classe, giorno, ora) ----
    records = []
    giorni_order, ore_order = [], []
    for _, row in df.iterrows():
        docente = tidy(row.get("Docente", ""))
        if not docente:
            continue
        for c in df.columns:
            if c == "Docente": continue
            if "_" not in c:   continue
            day, hour_s = c.rsplit("_", 1)
            try:
                hour = int(hour_s)
            except:
                continue
            cell = tidy(row.get(c, ""))
            if not cell:
                continue
            # la tabella sostegno contiene solo classi; scarta aule eventuali
            for tok in split_tokens(cell):
                if not tok: continue
                if _is_aula_token(tok, known_aule):
                    continue
                records.append({
                    "plesso":  "",           # sconosciuto: verrà usato solo per matching giorno/ora/classe
                    "docente": docente,
                    "classe":  tok,
                    "aula":    "",           # verrà riempita da integrate_sostegno_and_mark
                    "giorno":  day,
                    "ora":     hour,
                })
            if day not in giorni_order: giorni_order.append(day)
            if hour not in ore_order:   ore_order.append(hour)

    if not records:
        raise ValueError(f"{Path(tabella_sostegno_path).name}: nessuna classe riconosciuta nelle celle giorno/ora.")

    df_out = pd.DataFrame.from_records(records)
    ore_order = sorted(ore_order)
    df_out["giorno"] = pd.Categorical(df_out["giorno"], categories=giorni_order, ordered=True)
    df_out["ora"]    = pd.Categorical(df_out["ora"],    categories=ore_order,   ordered=True)
    return df_out, giorni_order, ore_order


def integrate_sostegno_and_mark(df_all, df_sostegno, df_aule=None, df_classi=None):
    """
    Integra in df_all i docenti di sostegno (df_sostegno: colonne plesso, docente, classe, aula="", giorno, ora)
    copiando l'aula dal docente titolare che ha la stessa classe nella stessa (giorno, ora) e
    marcando la classe con ^ sia al sostegno sia al titolare, preservando eventuali * finali.
    """
    import pandas as pd, re

    def tidy(x):
        if pd.isna(x): return ""
        s = str(x).strip()
        return "" if s.lower()=="nan" else s

    # stesso splitter del resto del codice
    def split_tokens(s):
        if not s: return []
        s = str(s).strip()
        for a,b in re.findall(r"\[(.*?)\]|\((.*?)\)", s):
            tok = (a or b or "").strip()
            if tok: s += f" | {tok}"
        return [p.strip() for p in re.split(r"\s*[\|/–—-]\s*", s) if p.strip()]

    # normalizzazione per confronto classe: togli ^, *, spazi multipli, lowercase
    def norm_for_match(tok: str) -> str:
        t = (tok or "").strip()
        t = re.sub(r"^\^+", "", t)       # via i caret iniziali
        t = re.sub(r"\*+$", "", t)       # via gli asterischi finali
        t = re.sub(r"\s+", " ", t)
        return t.lower().strip()

    # aggiunge un solo caret, preservando qualunque suffisso di * già presente
    def add_caret_preserving_star(tok: str) -> str:
        t = (tok or "").strip()
        # raccogli suffisso di asterischi
        m = re.search(r"\*+$", t)
        stars = m.group(0) if m else ""
        core = re.sub(r"\*+$", "", t)   # rimuovi gli * in coda
        core = core.lstrip("^")         # rimuovi eventuali ^ multipli
        return "^" + core + stars

    # applica il caret ai soli token la cui normalizzazione è in target_norms
    def add_caret_in_cell(cell: str, target_norms: set[str]) -> str:
        toks = split_tokens(cell)
        out = []
        for tok in toks:
            if norm_for_match(tok) in target_norms:
                out.append(add_caret_preserving_star(tok))
            else:
                out.append(tok)
        return " | ".join(out)

    # mappa aula->plesso (se disponibile) per assegnare il plesso al sostegno
    room2plesso = {}
    if df_aule is not None and {"Aula","Plesso"} <= set(df_aule.columns):
        for _, rr in df_aule.iterrows():
            a = tidy(rr.get("Aula","")).lower()
            p = tidy(rr.get("Plesso",""))
            if a:
                room2plesso[a] = p

    df_all = df_all.copy()
    # assicura colonna-flag per distinguere sostegno / non-sostegno
    has_flag = "is_sostegno" in df_all.columns
    df_all = df_all.copy()
    if not has_flag:
        df_all["is_sostegno"] = False

    new_rows = []

    # per velocizzare i match, pre-computo un indice (giorno,ora) -> indici righe df_all
    by_slot = {}
    for i, r in df_all.iterrows():
        key = (str(r["giorno"]), str(r["ora"]))
        by_slot.setdefault(key, []).append(i)

    for _, r in df_sostegno.iterrows():
        g = str(r["giorno"]); o = str(r["ora"])
        cls_tokens = [t for t in split_tokens(tidy(r["classe"])) if t]
        if not cls_tokens:
            continue
        target_norms = {norm_for_match(t) for t in cls_tokens}

        # cerca righe titolari nello stesso (giorno,ora) che contengono almeno una di queste classi
        idxs = by_slot.get((g, o), [])
        matched_idx = []
        aule_set = set()
        plessi_set = set()

        for i2 in idxs:
            cls_cell = tidy(df_all.at[i2, "classe"])
            if not cls_cell:
                continue
            toks2 = split_tokens(cls_cell)
            if any(norm_for_match(t2) in target_norms for t2 in toks2):
                matched_idx.append(i2)
                # accumula aule
                for a in split_tokens(tidy(df_all.at[i2, "aula"])):
                    if a:
                        aule_set.add(a)
                        p = room2plesso.get(a.lower())
                        if p: plessi_set.add(p)

        # aula assegnata al sostegno = unione (ordinata) delle aule trovate
        aula_val = " | ".join(sorted(aule_set))
        # plesso del sostegno: se unico tra quelli dedotti, usalo
        plesso_val = list(plessi_set)[0] if len(plessi_set) == 1 else tidy(r.get("plesso",""))

        # 1) aggiungi riga sostegno (classi marcate con ^, preservando *)
        sostegno_class = " | ".join(add_caret_preserving_star(t) for t in cls_tokens)
        new_rows.append({
            "plesso":       plesso_val,
            "docente":      tidy(r["docente"]),
            "classe":       sostegno_class,
            "aula":         aula_val,
            "giorno":       r["giorno"],
            "ora":          r["ora"],
            "is_sostegno":  True,   # <— flag esplicito
        })


        # 2) marca con ^ anche i docenti titolari corrispondenti (senza toccare eventuali *)
        for i2 in matched_idx:
            old_cell = tidy(df_all.at[i2, "classe"])
            df_all.at[i2, "classe"] = add_caret_in_cell(old_cell, target_norms)

    if new_rows:
        df_new = pd.DataFrame.from_records(new_rows)
        # mantieni le stesse categorie di giorno/ora se presenti
        if "giorno" in df_all and hasattr(df_all["giorno"], "cat"):
            df_new["giorno"] = pd.Categorical(df_new["giorno"],
                                              categories=list(df_all["giorno"].cat.categories),
                                              ordered=True)
        if "ora" in df_all and hasattr(df_all["ora"], "cat"):
            df_new["ora"] = pd.Categorical(df_new["ora"],
                                           categories=list(df_all["ora"].cat.categories),
                                           ordered=True)
        df_all = pd.concat([df_all, df_new], ignore_index=True)

    return df_all



def tidy(x):
    if pd.isna(x): return ""
    s = str(x).strip()
    return "" if s.lower() == "nan" else s

DAY_ALIASES = {
    "lun":"Lunedì","lunedi":"Lunedì","lunedì":"Lunedì",
    "mar":"Martedì","martedi":"Martedì","martedì":"Martedì",
    "mer":"Mercoledì","mercoledi":"Mercoledì","mercoledì":"Mercoledì",
    "gio":"Giovedì","giovedi":"Giovedì","giovedì":"Giovedì",
    "ven":"Venerdì","venerdi":"Venerdì","venerdì":"Venerdì",
    "sab":"Sabato","sabato":"Sabato",
    "dom":"Domenica","domenica":"Domenica",
}
def _norm_day(tok: str) -> str:
    t = (tok or "").strip().lower()
    return DAY_ALIASES.get(t, tok.strip())

def split_tokens(s):
    """Split robusto su | / – — - e porta in chiaro [..] e (..)."""
    if not s: return []
    s = str(s).strip()
    for a,b in re.findall(r"\[(.*?)\]|\((.*?)\)", s):
        tok = (a or b or "").strip()
        if tok: s += f" | {tok}"
    return [p.strip() for p in re.split(r"\s*[\|/–—-]\s*", s) if p.strip()]

def _norm_key_generic(s: str) -> str:
    s = tidy(s)
    s = re.sub(r"\[(.*?)\]|\((.*?)\)", " ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

# --- normalizzazione docente/classe (shared) ---
def _extract_docente_code(nome: str) -> str | None:
    """Codice tipo 'ZX_AS48' → 'zx_as48'."""
    s = tidy(nome)
    if not s: return None
    m = re.search(r"\b([A-Za-z]{2}_[A-Za-z0-9]{3,})\b", s)
    return m.group(1).lower() if m else None

def _strip_docente_tokens(nome: str) -> list[str]:
    """Tieni solo token alfabetici; scarta cifre, underscore, COE/PT/SUPP, parentesi."""
    s = tidy(nome)
    if not s: return []
    s = re.sub(r"\[(.*?)\]|\((.*?)\)", " ", s)
    s = s.replace("_", " ")
    tokens = re.split(r"[,\s;:/\-]+", s)
    out = []
    for t in tokens:
        t0 = t.strip()
        if not t0: continue
        if t0.upper().rstrip(".") in {"COE","PT","SUPP","SUP","POTENZIAMENTO"}:
            continue
        if any(ch.isdigit() for ch in t0):
            continue
        if re.fullmatch(r"[A-Za-zÀ-ÖØ-öø-ÿ]+", t0):
            out.append(t0)
    return out

def _norm_lookup_docente(nome: str) -> str:
    toks = _strip_docente_tokens(nome)
    if not toks: return ""
    full = " ".join(toks).lower().strip()
    return full if full else toks[-1].lower().strip()

def _norm_lookup_classe(classe: str) -> str:
    s = tidy(classe)
    # togli note tra parentesi
    s = re.sub(r"\[(.*?)\]|\((.*?)\)", " ", s)
    # >>> fix: rimuovi ^ iniziali e * finali
    s = re.sub(r"^\^+", "", s)
    s = re.sub(r"\*+$", "", s)
    # normalizza spazi e case
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


# ======= INTESTAZIONE GLOBALE (richiesta a video) =======

HEADER_TEXT: str | None = None  # usata da tutte le funzioni

def set_header_text(text: str | None):
    """Imposta l'intestazione globale"""
    global HEADER_TEXT
    HEADER_TEXT = (text or "").strip()
    if HEADER_TEXT:
        print(f"[OK] Intestazione impostata:\n{HEADER_TEXT}")
    else:
        print("[Avviso] Intestazione vuota: non verrà stampata.")

def get_header_text() -> str:
    """Ritorna l'intestazione corrente ('' se non impostata)."""
    return HEADER_TEXT or ""

def prompt_header_text(force: bool=False, default: str|None=None):
    """
    Chiede a video l'intestazione globale.
    - force=True: chiede sempre, anche se già impostata
    - default: valore proposto tra parentesi quadre
    """
    global HEADER_TEXT
    if not force and (HEADER_TEXT and HEADER_TEXT.strip()):
        # già presente: non chiedo
        print(f"[Info] Intestazione già impostata:\n{HEADER_TEXT}")
        return

    base_prompt = "Inserisci l'intestazione globale da stampare su Excel"
    if default:
        resp = input(f"{base_prompt} [{default}]: ").strip()
        text = resp if resp else default
    else:
        text = input(f"{base_prompt}: ").strip()
    set_header_text(text)

# ======= Callback helper Excel (invariati) =======
from openpyxl.styles import Alignment, Font, PatternFill


def xlsx_add_header(ws, ncols, text, align="right", bold=False, italic=True, fill=None):
    """Inserisce una riga di intestazione in alto, mergiata su ncols."""
    if not text:
        return
    ws.append([text])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1)
    c.font = Font(bold=bold, italic=italic)
    c.alignment = Alignment(horizontal={"left":"left","center":"center","right":"right"}[align],
                            vertical="center")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    ws.append([])  # riga vuota di separazione



# prompt_header_text(default="I.I.S. Via dei Papareschi - Orario Provvisorio valido dal 22/09 al 26/09")




# ================================
# Header builder tabelle giorno/ora
# ================================
def header_rows_for_day_hour(giorni_order, ore_order, first_col_title="Docente"):
    """Due righe (giorni mergiati + ore). Ritorna (rows, spans, day_bounds)."""
    row0 = [first_col_title]
    row1 = [""]
    spans = []
    day_bounds = []  # (start_col, end_col)
    col = 1
    for g in giorni_order:
        start = col
        for o in ore_order:
            row1.append(str(o)); col += 1
        end = col - 1
        row0 += [g] + [""]*(len(ore_order)-1)
        spans.append(("SPAN", (start,0), (end,0)))
        day_bounds.append((start, end))
    return [row0, row1], spans, day_bounds

def header_rows_for_day_hour_generic(giorni_order, ore_order):
    return header_rows_for_day_hour(giorni_order, ore_order, first_col_title="")





# ================================
# Parser orari (docenti 2 righe: classi/aule)
# ================================


def read_teacher_matrix(xlsx_path, sheet_name=0, plesso_label="", df_aule=None, max_trailer_rows=3):
    import re
    import pandas as pd
    from pathlib import Path

    # --- shim helpers: alias a quelli globali, con fallback ---
    def _tidy(x):
        try:
            return tidy(x)  # usa la tidy globale se presente
        except NameError:
            if pd.isna(x):
                return ""
            s = str(x).strip()
            return "" if s.lower() == "nan" else s

    try:
        _split_tokens = split_tokens  # alias al globale
    except NameError:
        def _split_tokens(s: str):
            if not s: return []
            s = str(s).strip()
            for a,b in re.findall(r"\[(.*?)\]|\((.*?)\)", s):
                tok = (a or b or "").strip()
                if tok: s += f" | {tok}"
            return [p.strip() for p in re.split(r"\s*[\|/–—-]\s*", s) if p.strip()]



    # -------------------- helpers locali --------------------

    DAY_ALIASES = {
        "lun":"Lunedì","lunedi":"Lunedì","lunedì":"Lunedì",
        "mar":"Martedì","martedi":"Martedì","martedì":"Martedì",
        "mer":"Mercoledì","mercoledi":"Mercoledì","mercoledì":"Mercoledì",
        "gio":"Giovedì","giovedi":"Giovedì","giovedì":"Giovedì",
        "ven":"Venerdì","venerdi":"Venerdì","venerdì":"Venerdì",
        "sab":"Sabato","sabato":"Sabato","dom":"Domenica","domenica":"Domenica",
    }
    def _norm_day(tok: str) -> str:
        t = (tok or "").strip().lower()
        return DAY_ALIASES.get(t, tok.strip())

    # aule note (migliora il riconoscimento dei token aula)
    known_aule = set()
    if df_aule is not None and "Aula" in df_aule.columns:
        known_aule |= {str(a).strip().lower() for a in df_aule["Aula"].dropna() if str(a).strip()}

    def _is_aula_token(tok: str) -> bool:
        t = (tok or "").strip().lower()
        if not t: return False
        if t in known_aule:
            return True
        # pattern generici (adatta se serve)
        return bool(re.match(r"^[A-Za-z]{0,4}\d{2,4}$", t)) or t.startswith("lab") or t.startswith("aula ")

    # -------------------- trova header e normalizza colonne --------------------
    df0 = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None, engine="openpyxl")
    header_row_idx = 0
    for i in range(min(10, len(df0))):
        v = _tidy(df0.iat[i,0]) if df0.shape[1] else ""
        if v.lower() == "docente":
            header_row_idx = i
            break

    # Prova lettura con DOPPIO header (gestisce header a due righe)
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=[header_row_idx, header_row_idx+1], engine="openpyxl")
    newcols, per_day_counter = [], {}
    for a,b in df.columns:
        a = _tidy(a); b = _tidy(b)
        day = _norm_day(a) if a else ""
        hour = ""
        if b:
            m = re.search(r"(\d+)", b)
            if m: hour = int(m.group(1))
        if day and hour == "":
            per_day_counter[day] = per_day_counter.get(day, 0) + 1
            hour = per_day_counter[day]
        if day and hour != "":
            newcols.append(f"{day}_{hour}")
        elif a:
            newcols.append(day)
        else:
            newcols.append(str(b))
    df.columns = [str(c).strip() for c in newcols]

    cols = df.columns.tolist()
    if not cols or cols[0].strip().lower() != "docente":
        # Piano B: header singolo
        df1 = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=header_row_idx, engine="openpyxl")
        df1.columns = [str(c).strip() for c in df1.columns]
        df, cols = df1, df1.columns.tolist()
    if not cols or cols[0].strip().lower() != "docente":
        raise ValueError(f"[{Path(xlsx_path).name}] prima colonna non è 'Docente' (trovato: '{cols[0] if cols else 'N/A'}').")

    # pulizia celle
    df[cols[0]] = df[cols[0]].apply(_tidy)
    for c in cols[1:]:
        df[c] = df[c].apply(_tidy)

    # individua colonne tempo Giorno_Ora -> (col, giorno, ora)
    parsed_cols, giorni_order, ore_order = [], [], []
    for c in cols[1:]:
        m = re.match(r"^([A-Za-zÀ-ÖØ-öø-ÿ]+)\s*[_\-\s]?\s*(\d+)$", str(c).strip())
        if not m:
            continue
        day = _norm_day(m.group(1))
        hour = int(m.group(2))
        parsed_cols.append((c, day, hour))
        if day not in giorni_order: giorni_order.append(day)
        if hour not in ore_order:   ore_order.append(hour)
    if not parsed_cols:
        raise ValueError(f"[{Path(xlsx_path).name}] colonne Giorno/Ora non riconosciute. Esempi: {cols[:8]}")

    # -------------------- percorri blocchi docente --------------------
    records = []
    i, n = 0, len(df)

    while i < n:
        docente = _tidy(df.iloc[i, 0])
        if not docente:
            i += 1
            continue

        # 1) CASO PRINCIPALE: la riga subito sotto ripete il docente = riga AUEL
        aule_row = None
        j = i + 1
        if j < n:
            next_first = _tidy(df.iloc[j, 0])
            if next_first and next_first.lower() == docente.lower():
                aule_row = df.iloc[j]
                j = i + 2  # salta anche la riga aule

        # 2) FALLBACK: cerca "trailer" con prima cella vuota o tag simili
        if aule_row is None:
            trailer = []
            j2 = i + 1

            def _tag_first(idx):
                return _tidy(df.iloc[idx, 0]).strip().lower()

            while j2 < n and len(trailer) < max_trailer_rows:
                tag = _tag_first(j2)
                # considera candidate aule se prima cella è vuota o indica 'aula'
                if tag in {"", "aula", "aule", "room", "rooms", "-", "aule/stanze"}:
                    trailer.append(df.iloc[j2])
                    j2 += 1
                else:
                    break

            # scegli come 'aule_row' quella che contiene più token aula nelle celle orarie
            best_score = -1
            best_row = None
            for cand in trailer:
                score = 0
                for col_name, _, _ in parsed_cols:
                    val = _tidy(cand[col_name])
                    if val and any(_is_aula_token(t) for t in _split_tokens(val)):
                        score += 1
                if score > best_score:
                    best_score = score
                    best_row = cand
            if best_row is not None:
                aule_row = best_row
                j = i + 1 + len(trailer)  # salta tutto il trailer
            else:
                j = i + 1  # nessuna riga aule riconosciuta

        # 3) Estrai celle orarie + fallback "Classe | Aula"
        classes_row = df.iloc[i]
        for col_name, g, o in parsed_cols:
            val_cls  = _tidy(classes_row[col_name])
            val_aula = _tidy(aule_row[col_name]) if aule_row is not None else ""

            # fallback: aula infilata nella cella di 'classe' (es. "5AS | C027")
            if not val_aula and val_cls:
                toks = _split_tokens(val_cls)
                room_toks = [t for t in toks if _is_aula_token(t)]
                if room_toks:
                    val_aula  = " | ".join(sorted(set(room_toks)))
                    class_toks = [t for t in toks if t not in room_toks]
                    val_cls   = " | ".join(class_toks)

            if val_cls:
                records.append({
                    "plesso":  plesso_label,
                    "docente": docente,
                    "classe":  val_cls,
                    "aula":    val_aula,
                    "giorno":  g,
                    "ora":     o
                })

        # 4) passa al prossimo docente
        i = max(j, i + 1)

    # -------------------- output ordinato --------------------
    long_df = pd.DataFrame.from_records(records)
    ore_order = sorted(ore_order)
    long_df["giorno"] = pd.Categorical(long_df["giorno"], categories=giorni_order, ordered=True)
    long_df["ora"]    = pd.Categorical(long_df["ora"],    categories=ore_order,   ordered=True)
    return long_df, giorni_order, ore_order


def load_tabella_materie(tabella_materie_path, sheet_name=0):
    """
    Costruisce materie_map: (docente_norm, classe_norm) -> MATERIA
    Gestisce il layout:
      MATERIA | DOCENTE | CLASSI | Unnamed:3 | Unnamed:4 | ...
    dove le CLASSI sono nelle celle (non nei nomi colonna).
    """
    import re
    import pandas as pd
    from pathlib import Path

    def tidy(x):
        if pd.isna(x): return ""
        s = str(x).strip()
        return "" if s.lower()=="nan" else s

    def _norm_key(s: str) -> str:
        s = tidy(s)
        s = re.sub(r"\[(.*?)\]|\((.*?)\)", " ", s)
        s = re.sub(r"\s+", " ", s).strip().lower()
        return s

    # --- leggi con header singolo auto-detect (riga 0..9) ---
    raw = pd.read_excel(tabella_materie_path, sheet_name=sheet_name, header=None, dtype=str, engine="openpyxl")

    header_idx = None
    for i in range(min(10, len(raw))):
        cols = [str(c).strip().lower() for c in pd.read_excel(tabella_materie_path, sheet_name=sheet_name,
                                                              header=i, nrows=0, engine="openpyxl").columns]
        if "materia" in cols and "docente" in cols:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"{Path(tabella_materie_path).name}: non trovo header con 'Materia' e 'Docente'.")

    df = pd.read_excel(tabella_materie_path, sheet_name=sheet_name, header=header_idx, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    # individua nomi colonne chiave (case-insensitive)
    def find_col(name):
        for c in df.columns:
            if c.strip().lower() == name:
                return c
        return None

    col_materia = find_col("materia")
    col_docente = find_col("docente")
    col_classi  = find_col("classi")  # prima colonna delle classi

    if col_materia is None or col_docente is None or col_classi is None:
        raise ValueError(f"{Path(tabella_materie_path).name}: servono 'Materia', 'Docente' e 'Classi'.")

    # tutte le colonne "di classe" sono da 'CLASSI' in poi, incluse le Unnamed
    start_idx = list(df.columns).index(col_classi)
    class_cols = list(df.columns)[start_idx:]

    # build map
    materie_map = {}
    rows_preview = []  # per debug opzionale

    for _, r in df.iterrows():
        materia = tidy(r.get(col_materia, ""))      # es. "A012 DISC.LETT."
        docente = tidy(r.get(col_docente, ""))      # es. "MATTEINI" o "ZX_BB02"
        if not materia or not docente:
            continue

        d_norm  = _norm_lookup_docente(docente)     # normalizzazione robusta (già definita nel tuo codice)
        d_code  = _extract_docente_code(docente)    # "zx_bb02" se presente
        d_raw   = docente.replace("_"," ").lower().strip()

        # scorri tutte le colonne classi: prendi le CELLE non vuote come nomi classe
        for cc in class_cols:
            cls_val = tidy(r.get(cc, ""))
            if not cls_val:
                continue
            # la cella può contenere anche più classi separate da spazi/virgole? (gestiamo i separatori più comuni)
            parts = re.split(r"[;,/\s]+", cls_val)
            for part in parts:
                part = part.strip()
                if not part:
                    continue
                c_norm = _norm_lookup_classe(part)

                # registra 3 chiavi alternative per massimizzare i match
                if d_norm:
                    materie_map[(d_norm, c_norm)] = materia
                if d_code:
                    materie_map[(d_code, c_norm)] = materia
                if d_raw:
                    materie_map[(d_raw,  c_norm)] = materia

                # facoltativo per ispezione
                if len(rows_preview) < 20:
                    rows_preview.append({"docente_norm": d_norm or d_code or d_raw,
                                         "classe_norm": c_norm, "materia": materia})

    # DEBUG: prime 20 righe “esplicitate” (se vuoi)
    if DEBUG:
        print(f"[load_tabella_materie] Associazioni create: {len(materie_map)}")
        if rows_preview:
            print("[load_tabella_materie] Esempi (prime 20):")
            for i, r in enumerate(rows_preview):
                print(f"  {i+1:02d} -> ({r['docente_norm']}, {r['classe_norm']}) => {r['materia']}")
    return materie_map



# ================================
# EXPORT 1: Orario Aule e Classe Settimanale (filtrato da Tabella_Classi)
# ================================

def export_OUTPUT_AULE_SETTIMANALE(
    df_all,
    df_aule,
    titolo="ORARIO_AULE_SETTIMANALE",
    plesso=None,
    # ======= XLSX: larghezze colonne (caratteri) =======
    xlsx_first_col_width=6.5,
    xlsx_second_col_width=10.0,
    xlsx_day_col_width=None,
    xlsx_day_col_min=14.0,
    xlsx_day_col_max=60.0,
):
    import re, openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # ---------- dati ----------
    if df_aule is None or df_aule.empty:
        raise ValueError("df_aule è vuoto: caricare Tabella_Aule.")

    known_aule = {str(a).strip().lower() for a in df_aule["Aula"].dropna() if str(a).strip()}

    df_aule_ord = df_aule.copy()
    if plesso is not None and "Plesso" in df_aule_ord.columns:
        df_aule_ord = df_aule_ord[
            df_aule_ord["Plesso"].astype(str).str.strip().str.lower()
            == str(plesso).strip().lower()
        ]
    allowed_aule = [a for a in df_aule_ord["Aula"].astype(str).tolist() if a and a.strip()]
    allowed_aule_ci = {a.strip().lower() for a in allowed_aule}

    def row_has_allowed_aula(aula_str, classe_str):
        for t in split_tokens(aula_str):
            if t.lower() in allowed_aule_ci:
                return True
        for t in split_tokens(classe_str):
            if t.lower() in allowed_aule_ci:
                return True
        return False

    df_only = df_all[df_all.apply(
        lambda r: row_has_allowed_aula(str(r.get("aula","")), str(r.get("classe",""))),
        axis=1
    )].copy()

    def clean_aula(aula_str, classe_str):
        toks = [t for t in split_tokens(aula_str) if t.lower() in allowed_aule_ci]
        if not toks:
            toks = [t for t in split_tokens(classe_str) if t.lower() in allowed_aule_ci]
        return " | ".join(sorted(set(toks)))
    df_only["aula"] = df_only.apply(
        lambda r: clean_aula(str(r.get("aula","")), str(r.get("classe",""))),
        axis=1
    )

    present_ci = {t.strip().lower() for s in df_only["aula"].astype(str)
                  for t in split_tokens(s) if t.strip()}
    aule_order = [a for a in allowed_aule if a.strip().lower() in present_ci]
    if not aule_order:
        raise ValueError("Nessuna delle aule in Tabella_Aule è presente nei dati dell'orario.")

    # Assi giorni/ore robusti anche se le colonne non sono 'category'
    try:
        giorni = list(df_only["giorno"].cat.categories)
    except (AttributeError, TypeError):
        # mantieni l'ordine di apparizione
        giorni = list(dict.fromkeys(df_only["giorno"].astype(str).tolist()))

    try:
        ore = list(df_only["ora"].cat.categories)
    except (AttributeError, TypeError):
        ore_vals = [str(x) for x in df_only["ora"].tolist()]
        # se tutte numeriche -> ordinale; altrimenti ordine di apparizione
        if ore_vals and all(v.isdigit() for v in ore_vals):
            ore = sorted(set(ore_vals), key=lambda v: int(v))
        else:
            ore = list(dict.fromkeys(ore_vals))


    # ---------- Excel ----------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aule"
    ncols = 2 + len(giorni)

    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_color = "D9E3F0"
    try:
        hdr_color = XLSX(COLOR_HEADER_HEX)
    except Exception:
        pass
    fill_title  = PatternFill("solid", fgColor=hdr_color)
    fill_header = PatternFill("solid", fgColor=hdr_color)

    def _append(values):
        ws.append(values)

    def _apply_border_range(r1, c1, r2, c2):
        for rr in range(r1, r2+1):
            for cc in range(c1, c2+1):
                ws.cell(row=rr, column=cc).border = border_all

    def _append_aula_title(aula_name: str):
        """Riga titolo (prime 3 celle mergiate)."""
        start_row = ws.max_row + 1
        _append([""] * ncols)  # riga placeholder
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=min(3, ncols))
        tcell = ws.cell(row=start_row, column=1)
        title_text = f"Orario settimanale – Aula {str(aula_name).strip()}"
        if title_text.startswith("="):
            title_text = "'" + title_text
        tcell.value = title_text
        tcell.font = Font(bold=True)
        tcell.fill = fill_title
        tcell.alignment = Alignment(horizontal="left", vertical="center")
        end_c = min(3, ncols)
        for c in range(1, end_c + 1):
            ws.cell(row=start_row, column=c).border = Border(
                left=thin if c == 1 else None,
                right=thin if c == end_c else None,
                top=thin,
                bottom=thin,
            )
        return start_row

    # ===== Intestazione globale UNA SOLA VOLTA in alto =====
    excel_add_global_header(ws, ncols)

    # colora le prime 2 righe come header del progetto
    for r in (1, 2):
        for cc in range(1, ncols+1):
            ws.cell(row=r, column=cc).fill = fill_header

    # ===== LOOP AULE =====
    for aula in aule_order:
        # titoletto
        _append_aula_title(aula)

        # subset per l'aula
        pat = rf"(?:^|\s*\|\s*){re.escape(aula.lower())}(?:\s*\|\s*|$)"
        sub = df_only[df_only["aula"].str.lower().str.contains(pat, regex=True, na=False)]

        # header tabella (una riga)
        header = ["Ora", ""] + giorni
        header_row = ws.max_row + 1
        _append(header)
        for cc in range(1, ncols+1):
            ws.cell(row=header_row, column=cc).fill = fill_header
            ws.cell(row=header_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")

        # corpo: 2 righe per ora (Docente / Classe)
        for o in ore:
            start_time = ORE_MAP.get(int(o), str(o)) if isinstance(o, (int, float, str)) else str(o)
            row_doc = [start_time, "Docente"]
            row_cls = ["",         "Classe"]

            for g in giorni:
                cell = sub[(sub["giorno"] == g) & (sub["ora"] == o)]

                # docenti
                docenti = sorted({tidy(x) for x in cell["docente"].tolist() if tidy(x)})
                doc_txt = " | ".join(docenti)

                # classi: escludi token aula e pulisci ^ / *
                class_set = set()
                for raw in cell["classe"].astype(str).tolist():
                    for t in split_tokens(raw):
                        if t.lower() in known_aule:
                            continue
                        t2 = norm_class_token(t)
                        if t2:
                            class_set.add(t2)
                cls_txt = " | ".join(sorted(class_set))

                row_doc.append(doc_txt)
                row_cls.append(cls_txt)

            # scrittura + merge "Ora"
            _append(row_doc); _append(row_cls)
            last_row = ws.max_row
            ws.merge_cells(start_row=last_row-1, start_column=1, end_row=last_row, end_column=1)
            ws.cell(row=last_row-1, column=1).alignment = Alignment(horizontal="center", vertical="center")

            # stile: Docente bold, Classe italic
            for cc in range(3, ncols+1):
                ws.cell(row=last_row-1, column=cc).font = Font(bold=True)     # Docente
                ws.cell(row=last_row,   column=cc).font = Font(italic=True)   # Classe

        # bordi/align blocco tabella
        end_row_block = ws.max_row
        _apply_border_range(header_row, 1, end_row_block, ncols)
        for rr in range(header_row, end_row_block+1):
            ws.cell(row=rr, column=2).alignment = Alignment(horizontal="left", vertical="center")
            for cc in range(3, ncols+1):
                ws.cell(row=rr, column=cc).alignment = Alignment(
                    wrap_text=True, vertical="center", horizontal="center"
                )

        # riga vuota di separazione tra un'aula e la successiva
        ws.append([""] * ncols)

    # ---------- Excel: larghezze ----------
    total_cols = 2 + len(giorni)
    ws.column_dimensions[get_column_letter(1)].width = float(xlsx_first_col_width)
    ws.column_dimensions[get_column_letter(2)].width = float(xlsx_second_col_width)
    if xlsx_day_col_width is None:
        max_chars_days = 0
        for row in ws.iter_rows(min_col=3, max_col=total_cols, values_only=True):
            for val in row:
                if val is None:
                    continue
                max_chars_days = max(max_chars_days, len(str(val)))
        auto_w = max(xlsx_day_col_min, min(xlsx_day_col_max, max_chars_days + 2))
        day_xlsx_w = float(auto_w)
    else:
        day_xlsx_w = float(xlsx_day_col_width)
    for col_idx in range(3, total_cols+1):
        ws.column_dimensions[get_column_letter(col_idx)].width = day_xlsx_w

    xlsx_path = OUTPUT_DIR / f"{titolo}.xlsx"
    wb.save(xlsx_path)

    print("Creato XLSX:", xlsx_path)
    return xlsx_path





def export_OUTPUT_CLASSI_SETTIMANALE(
    df_all,
    df_classi,
    titolo="ORARIO_CLASSI_SETTIMANALE",
    materie_map=None,
    df_aule=None,                          # opzionale: per riconoscere token aula in fallback
    # ======= XLSX: larghezze colonne (unità Excel ~ caratteri) =======
    xlsx_first_col_width=6.5,              # "Ora"
    xlsx_second_col_width=10.0,            # "Docente/Materia/Aula"
    xlsx_day_col_width=None,               # se None => auto; altrimenti fisso per tutte le colonne dei giorni
    xlsx_day_col_min=14.0,
    xlsx_day_col_max=60.0,
):
    import re, openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    fill_header, fill_title, _ = excel_get_fills()

    # ---------------- Precondizioni / setup ----------------
    if df_classi is None or df_classi.empty:
        raise ValueError("df_classi è vuoto: caricare Tabella_Classi.")
    if materie_map is None:
        materie_map = {}

    def _extract_docente_code(nome: str):
        s = tidy(nome)
        if not s: return None
        m = re.search(r"\b([A-Za-z]{2}_[A-Za-z0-9]{3,})\b", s)
        return m.group(1).lower() if m else None

    def _strip_docente_tokens(nome: str):
        s = tidy(nome)
        if not s: return []
        s = re.sub(r"\[(.*?)\]|\((.*?)\)", " ", s)
        s = s.replace("_", " ")
        tokens = re.split(r"[,\s;:/\-]+", s)
        out = []
        for t in tokens:
            t0 = t.strip()
            if not t0: continue
            if t0.upper().rstrip(".") in {"COE","PT","SUPP","SUP","POTENZIAMENTO"}:
                continue
            if any(ch.isdigit() for ch in t0):
                continue
            if re.fullmatch(r"[A-Za-zÀ-ÖØ-öø-ÿ]+", t0):
                out.append(t0)
        return out

    def _norm_lookup_docente(nome: str) -> str:
        toks = _strip_docente_tokens(nome)
        if not toks: return ""
        full = " ".join(toks).lower().strip()
        return full if full else toks[-1].lower().strip()

    def _get_indirizzo_for(cls_name: str) -> str:
        if "Indirizzo" in df_classi.columns:
            s = df_classi.loc[df_classi["Classe"].astype(str)==cls_name, "Indirizzo"]
            if not s.empty: return str(s.iloc[0])
        return ""

    # aule note (per distinguere eventuali token aula nella colonna 'classe')
    known_aule = {str(a).strip().lower() for a in df_all["aula"].dropna() if str(a).strip()}
    if df_aule is not None and "Aula" in df_aule.columns:
        known_aule |= {str(a).strip().lower() for a in df_aule["Aula"].dropna() if str(a).strip()}

    # ===== Ordine classi: esattamente come in Tabella_Classi =====
    allowed_classes = [c for c in df_classi["Classe"].astype(str).tolist() if c and c.strip()]
    allowed_ci = [_norm_lookup_classe(c) for c in allowed_classes]
    allowed_set_ci = set(allowed_ci)

    # Filtro df_all a sole righe che appartengono a classi della tabella (ignorando eventuali token aula)
    def is_allowed_class(raw_cls):
        toks = [t for t in split_tokens(raw_cls) if t.lower() not in known_aule]
        if not toks:
            return False
        return any(_norm_lookup_classe(t) in allowed_set_ci for t in toks)

    df_only = df_all[df_all["classe"].astype(str).apply(is_allowed_class)].copy()

    # pulizia 'Classe | Aula' -> lascia solo classi in colonna 'classe' (normalizzate)
    def clean_class(raw_cls):
        tokens = [t for t in split_tokens(raw_cls) if t.lower() not in known_aule]
        keep = [_norm_lookup_classe(t) for t in tokens if _norm_lookup_classe(t) in allowed_set_ci]
        return " | ".join(keep) if keep else ""

    df_only["classe"] = df_only["classe"].astype(str).apply(clean_class)

    # prendi solo classi che compaiono, rispettando l'ordine di Tabella_Classi
    present_ci = {c.strip().lower() for c in df_only["classe"].unique() if str(c).strip()}
    classi_order = [c for c in allowed_classes if _norm_lookup_classe(c) in present_ci]
    if not classi_order:
        raise ValueError("Nessuna delle classi in Tabella_Classi è presente nei dati dell'orario.")

    # asse giorni/ore
    giorni = list(df_only["giorno"].cat.categories)
    ore    = list(df_only["ora"].cat.categories)

    def _texts_for_cell(sub_df, cls_name, g, o):
        # docenti
        docenti = sorted({tidy(x) for x in sub_df["docente"].tolist() if tidy(x)})
        doc_txt = " | ".join(docenti)
        # aule: colonna 'aula' + fallback da 'classe'
        aule_set = {tidy(x) for x in sub_df["aula"].tolist() if tidy(x)}
        for raw in sub_df["classe"].tolist():
            for tok in split_tokens(raw):
                if tok.lower() in known_aule:
                    aule_set.add(tok)
        aul_txt = " | ".join(sorted({a for a in aule_set if a}))
        # materie
        materie_set = set()
        classe_norm = _norm_lookup_classe(cls_name)
        for d in docenti:
            d_norm = _norm_lookup_docente(d)
            if d_norm:
                m = materie_map.get((d_norm, classe_norm))
                if m: materie_set.add(m); continue
            d_code = _extract_docente_code(d)
            if d_code:
                m = materie_map.get((d_code, classe_norm))
                if m: materie_set.add(m); continue
            d_raw = tidy(d).replace("_"," ").lower().strip()
            if d_raw:
                m = materie_map.get((d_raw, classe_norm))
                if m: materie_set.add(m)
        mat_txt = " | ".join(sorted(materie_set))
        return doc_txt, mat_txt, aul_txt

    # ---------------- Excel (unico foglio) ----------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ncols = 2 + len(giorni)
    ws.title = "Classi"

    # stili
    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    # usa il colore header del progetto, se presente; altrimenti un azzurrino di fallback
    hdr_color = "D9E3F0"
    try:
        hdr_color = XLSX(COLOR_HEADER_HEX)
    except Exception:
        pass
    fill_title  = PatternFill("solid", fgColor=hdr_color)
    fill_header = PatternFill("solid", fgColor=hdr_color)

    def _append(values):
        ws.append(values)

    def _apply_border_range(r1, c1, r2, c2):
        for rr in range(r1, r2+1):
            for cc in range(c1, c2+1):
                ws.cell(row=rr, column=cc).border = border_all

    def _append_class_title(cls_name: str):
        """Riga titolo (prime 3 celle mergiate)."""
        start_row = ws.max_row + 1
        _append([""] * ncols)  # riga placeholder
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=min(3, ncols))
        tcell = ws.cell(row=start_row, column=1)
        indirizzo = _get_indirizzo_for(cls_name)
        txt = f"Orario settimanale – Classe {str(cls_name).strip()}" + (f" — {indirizzo}" if indirizzo else "")
        if txt.startswith("="):
            txt = "'" + txt
        tcell.value = txt
        tcell.font = Font(bold=True)
        tcell.fill = fill_title
        tcell.alignment = Alignment(horizontal="left", vertical="center")
        end_c = min(3, ncols)
        for c in range(1, end_c + 1):
            ws.cell(row=start_row, column=c).border = Border(
                left=thin if c == 1 else None,
                right=thin if c == end_c else None,
                top=thin,
                bottom=thin,
            )
        return start_row

    # ===== Intestazione globale UNA SOLA VOLTA in alto =====
    excel_add_global_header(ws, ncols)

    # colora le prime 2 righe come header del progetto
    for r in (1, 2):
        for cc in range(1, ncols+1):
            ws.cell(row=r, column=cc).fill = fill_header

    # ===== LOOP CLASSI =====
    for cls in classi_order:
        # titoletto
        _append_class_title(cls)

        # header tabella (una riga)
        header_top = ["Ora", ""] + giorni
        header_row = ws.max_row + 1
        _append(header_top)
        for cc in range(1, ncols+1):
            ws.cell(row=header_row, column=cc).fill = fill_header
            ws.cell(row=header_row, column=cc).alignment = Alignment(horizontal="center", vertical="center")

        # subset di righe per la classe
        sub = df_only[df_only["classe"].str.lower() == _norm_lookup_classe(cls)]

        # corpo: 3 righe per ora
        for o in ore:
            start_time = ORE_MAP.get(int(o), str(o)) if isinstance(o, (int, float, str)) else str(o)
            row_doc = [start_time, "Docente"]
            row_mat = ["",         "Materia"]
            row_aul = ["",         "Aula"]

            for g in giorni:
                cell = sub[(sub["giorno"] == g) & (sub["ora"] == o)]
                doc_txt, mat_txt, aul_txt = _texts_for_cell(cell, cls, g, o)
                row_doc.append(doc_txt)
                row_mat.append(mat_txt)
                row_aul.append(aul_txt)

            # scrivi le 3 righe + merge "Ora"
            _append(row_doc); _append(row_mat); _append(row_aul)
            last_row = ws.max_row
            ws.merge_cells(start_row=last_row-2, start_column=1, end_row=last_row, end_column=1)
            ws.cell(row=last_row-2, column=1).alignment = Alignment(horizontal="center", vertical="center")

            # stile: Docente bold, Aula italic
            for cc in range(3, ncols+1):
                ws.cell(row=last_row-2, column=cc).font = Font(bold=True)      # Docente
                ws.cell(row=last_row-1, column=cc).font = Font()               # Materia
                ws.cell(row=last_row,   column=cc).font = Font(italic=True)    # Aula

        # bordi/align sul blocco tabella (da header_row all'ultima riga scritta)
        end_row_block = ws.max_row
        _apply_border_range(header_row, 1, end_row_block, ncols)
        for rr in range(header_row, end_row_block+1):
            ws.cell(row=rr, column=2).alignment = Alignment(horizontal="left", vertical="center")
            for cc in range(3, ncols+1):
                ws.cell(row=rr, column=cc).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

        # === riga vuota di separazione tra una classe e la successiva ===
        ws.append([""] * ncols)

    # ---------------- Excel: larghezze ----------------
    total_cols = ncols
    ws.column_dimensions[get_column_letter(1)].width = float(xlsx_first_col_width)
    ws.column_dimensions[get_column_letter(2)].width = float(xlsx_second_col_width)
    if xlsx_day_col_width is None:
        max_chars_days = 0
        for row in ws.iter_rows(min_col=3, max_col=total_cols, values_only=True):
            for val in row:
                if val is None:
                    continue
                ln = len(str(val))
                if ln > max_chars_days: max_chars_days = ln
        auto_w = max(xlsx_day_col_min, min(xlsx_day_col_max, max_chars_days + 2))
        day_xlsx_w = float(auto_w)
    else:
        day_xlsx_w = float(xlsx_day_col_width)
    for col_idx in range(3, total_cols+1):
        ws.column_dimensions[get_column_letter(col_idx)].width = day_xlsx_w

    # ---------------- Salvataggio Excel ----------------
    xlsx_path = OUTPUT_DIR / f"{titolo}.xlsx"
    wb.save(xlsx_path)

    print("Creato XLSX:", xlsx_path)
    return xlsx_path






# ================================
# EXPORT 2: Tabella Globale (A3, docenti su due righe, ordine alfabetico)
# ================================
def export_OUTPUT_TABELLA_GLOBALE(
    df_all, giorni, ore, df_aule=None,
    docenti_set=None,   # opzionale: set di docenti "validi" (per la compresenza)
    df_classi=None,
):

    import re
    from collections import defaultdict
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # ---------- helper locali ----------
    def tidy(x):
        import pandas as pd
        if pd.isna(x): return ""
        s = str(x).strip()
        return "" if s.lower() == "nan" else s

    def split_tokens(s):
        if not s: return []
        s = str(s).strip()
        # porta in chiaro [..] e (..)
        for a,b in re.findall(r"\[(.*?)\]|\((.*?)\)", s):
            tok = (a or b or "").strip()
            if tok: s += f" | {tok}"
        return [p.strip() for p in re.split(r"\s*[\|/–—-]\s*", s) if p.strip()]

    _norm_classe = _norm_lookup_classe

    def _is_aula(tok: str, known_aule) -> bool:
        t = (tok or "").strip().lower()
        if not t: return False
        if t in known_aule: return True
        return bool(re.match(r"^[A-Za-z]{0,4}\d{2,4}$", t)) or t.startswith("lab") or t.startswith("aula ")



    # --- NUOVO: mappa classe_norm -> tag plesso 'C'/'S' (da Tabella_Classi) ---
    def _build_class_plesso_tag(df_classi):
        mp = {}
        if df_classi is None:
            return mp
        # Nel tuo loader le colonne sono ["Edificio","Classe"]; se hai "Plesso", gestiamo anche quello
        pl_col = "Edificio" if "Edificio" in df_classi.columns else ("Plesso" if "Plesso" in df_classi.columns else None)
        if pl_col is None or "Classe" not in df_classi.columns:
            return mp
        for _, r in df_classi.iterrows():
            cls = _norm_lookup_classe(tidy(r.get("Classe","")))
            pl  = tidy(r.get(pl_col,"")).lower()
            if not cls or not pl:
                continue
            if "centr" in pl:
                mp[cls] = "C"
            elif "succ" in pl:
                mp[cls] = "S"
        return mp

    class_pl_tag = _build_class_plesso_tag(df_classi)

    # --- helper: classifica la cella CLASSI in 'C'/'S'/None ---
    def _classify_cell_by_classes(cls_cell: str, known_aule) -> str | None:
        # 1) prova a dedurre dal plesso delle classi (ignorando token-aula)
        tags = set()
        for t in split_tokens(tidy(cls_cell)):
            if _is_aula(t, known_aule):
                continue
            cn = _norm_classe(t)  # alias già definito: _norm_classe = _norm_lookup_classe
            if not cn:
                continue
            tg = class_pl_tag.get(cn)
            if tg:
                tags.add(tg)

        if tags == {"C"}:
            return "C"
        if tags == {"S"}:
            return "S"

        # 2) fallback: token di disposizione D (centrale) / d (succursale)
        dd = _dispo_tag_from_classes_cell(cls_cell)
        if dd:
            return dd

        return None


    # set delle classi "valide" tratte da Tabella_Classi (normalizzate)
    classes_valid_set = None
    if df_classi is not None and "Classe" in df_classi.columns:
        classes_valid_set = {
           _norm_classe(str(c))
            for c in df_classi["Classe"].dropna().astype(str)
            if _norm_classe(str(c))
        }


    # set docenti "validi" per compresenza
    if docenti_set is None:
        docenti_set = {tidy(d) for d in df_all["docente"].astype(str).dropna() if tidy(d)}

    # aule note
    known_aule = {str(a).strip().lower() for a in df_all["aula"].dropna() if str(a).strip()}
    if df_aule is not None and "Aula" in df_aule.columns:
        known_aule |= {str(a).strip().lower() for a in df_aule["Aula"].dropna() if str(a).strip()}

    # mappa Aula -> Plesso ('C' o 'S')
    room2plesso = {}
    if df_aule is not None and {"Aula","Plesso"} <= set(df_aule.columns):
        for _, r in df_aule.iterrows():
            a = tidy(r.get("Aula", ""))
            p = tidy(r.get("Plesso", "")).lower()
            if not a: continue
            tag = "C" if "centr" in p else ("S" if "succ" in p else None)
            room2plesso[a.lower()] = tag

    # header (giorni mergiati + ore)
    header, spans, day_bounds = header_rows_for_day_hour(giorni, ore, first_col_title="Docente")

    # ---------- mapping base per celle ----------
    # (docente,giorno,ora) -> classi/aule e plesso-set della cella
    grp = df_all.groupby(["docente","giorno","ora"], dropna=False, observed=False)
    classes_map, aule_map = defaultdict(list), defaultdict(list)
    cell_plessi_map = {}  # (d,g,o) -> set di plessi string ("Centrale"/"Succursale") presenti nelle righe sorgente

    for (d,g,o), sub in grp:
        key = (str(d), str(g), str(o))
        # classi / aule (pulendo "aula" da 'classe' se necessario)
        for raw in sub["classe"].dropna().tolist():
            for tok in split_tokens(raw):
                (aule_map if _is_aula(tok, known_aule) else classes_map)[key].append(tok)
        for raw in sub["aula"].dropna().tolist():
            for tok in split_tokens(raw):
                aule_map[key].append(tok)
        cell_plessi_map[key] = {tidy(p) for p in sub["plesso"].astype(str).unique() if tidy(p)}

    classes_map = {k: " | ".join(sorted(set(v))) for k,v in classes_map.items()}
    aule_map    = {k: " | ".join(sorted(set(v))) for k,v in aule_map.items()}

    # ---------- PRECALCOLO COMPRESENZE ----------
    # Regola: due docenti NON-sostegno, stessa classe, stessa ora, STESSA AULA
    # chiave: (giorno, ora, classe_norm, aula_norm) -> set(docenti_non_sostegno)
    comp_map = defaultdict(set)

    def _is_aula_token(tok: str) -> bool:
        return _is_aula(tok, known_aule)


    for _, r in df_all.iterrows():
        pl = tidy(r.get("plesso", ""))  # non più usato per la chiave
        g  = tidy(r.get("giorno", ""))
        o  = tidy(r.get("ora", ""))
        d  = tidy(r.get("docente", ""))
        if not (g and o and d):
            continue
        # escludi sostegno
        if str(r.get("is_sostegno", "")).strip().lower() in {"true","1"}:
            continue

        # classi (solo token-classe)
        class_norms = []
        for tok in split_tokens(tidy(r.get("classe",""))):
            if _is_aula_token(tok):
                continue
            cn = _norm_classe(tok)
            if not cn:
                continue
            if classes_valid_set is not None and cn not in classes_valid_set:
                continue
            class_norms.append(cn)

        if not class_norms:
            continue

        # aule candidate = token in colonna 'aula' + eventuali token-aula nella colonna 'classe'
        aule_tokens = set()
        for tok in split_tokens(tidy(r.get("aula",""))):
            if _is_aula_token(tok): aule_tokens.add(tok.strip().lower())
        for tok in split_tokens(tidy(r.get("classe",""))):
            if _is_aula_token(tok): aule_tokens.add(tok.strip().lower())

        # se non c'è aula non può esserci l'asterisco (richiede stessa aula)
        if not aule_tokens:
            continue

        for cn in class_norms:
            for aula_norm in aule_tokens:
                comp_map[(str(g), str(o), cn, aula_norm)].add(d)

    # chiavi dove ci sono almeno 2 docenti non-sostegno nella stessa aula
    comp_keys = {k for k, ds in comp_map.items() if len(ds) >= 2}



    # docenti unici ordinati
    docenti = sorted({d.strip() for d in df_all["docente"].astype(str).dropna() if d.strip()},
                     key=lambda x: x.lower())

    # helper PDF: classifica una cella aule in 'C', 'S' o None (per evidenziare Succursale)
    # --------- Classifica cella -> 'C' / 'S' / None (usa D/d, poi aula, poi fallback sorgente) ---------
    def classify_plesso_cell(key, aule_cell: str, cls_cell: str) -> str | None:
        # 1) priorità a D/d presenti nella cella CLASSI
        dd = _dispo_tag_from_classes_cell(cls_cell)
        if dd in {"C", "S"}:
            return dd

        # 2) prova a dedurre dal plesso delle AULE
        s = tidy(aule_cell)
        tags = set()
        if s:
            for t in [t.strip() for t in s.split("|") if t.strip()]:
                tag = room2plesso.get(t.lower())
                if tag in {"C","S"}:
                    tags.add(tag)
        if tags == {"C"}: return "C"
        if tags == {"S"}: return "S"

        # 3) fallback: plessi sorgente (se univoci)
        pl_full_set = cell_plessi_map.get(key, set())
        if len(pl_full_set) == 1:
            only = next(iter(pl_full_set)).lower()
            if "centr" in only: return "C"
            if "succ"  in only: return "S"

        return None



    # ========== Costruzione matrice + coordinate evidenziazione ==========
    data = [header[0], header[1]]
    start_body = len(data)
    n_hours = len(ore)
    n_time_cols = len(giorni) * len(ore)

    per_teacher_plesso_tags = [] # per Excel: lista tag 'C'/'S'/None per ogni colonna tempo, per riga CLASSI

    for d in docenti:
        r_cls = [d]   # riga CLASSI
        r_au  = [""]  # riga AULE
        pl_tags = []

        for g in giorni:
            for o in ore:
                k = (str(d), str(g), str(o))
                cls_txt = classes_map.get(k, "")
                aul_txt = aule_map.get(k, "")

                # ---- COMPRESENZA: aggiungi * ai token classe che risultano in compresenza ----
                if cls_txt:
                    toks = split_tokens(cls_txt)
                    # aule effettive della cella corrente
                    aule_cell = set()
                    for tok_a in split_tokens(aul_txt):
                        if _is_aula(tok_a, known_aule): aule_cell.add(tok_a.strip().lower())
                    # aggiungi eventuali token-aula presenti (per errore) nella colonna classi
                    for tok_a in split_tokens(classes_map.get(k, "")):
                        if _is_aula(tok_a, known_aule): aule_cell.add(tok_a.strip().lower())


                    new_toks = []
                    for t in toks:
                        cn = _norm_classe(t)
                        if classes_valid_set is not None and cn not in classes_valid_set:
                            new_toks.append(t); continue
                        starred = any((str(g), str(o), cn, a_norm) in comp_keys for a_norm in aule_cell)
                        new_toks.append(re.sub(r"\*+$", "", t) + ("*" if starred else ""))
                    cls_txt = " | ".join(new_toks)



                r_cls.append(cls_txt)
                r_au.append(aul_txt)

                # tag per evidenziare Succursale in PDF/Excel
                pl_tags.append(_classify_cell_by_classes(cls_txt, known_aule))

        # --- Logica "t": marca nella riga CLASSI la cella vuota prima del cambio plesso ---
        for di, g in enumerate(giorni):
            base = di * n_hours
            last_tag = None
            for h in range(n_hours):
                cur_tag = pl_tags[base + h]
                if cur_tag in {"C","S"}:
                    if last_tag is None:
                        last_tag = cur_tag
                    elif cur_tag != last_tag:
                        j = h
                        if j - 1 >= 0:
                            target_idx = 1 + base + (j - 1)  # +1 per 'Docente'
                            if tidy(r_cls[target_idx]) == "":
                                r_cls[target_idx] = "t"
                        last_tag = cur_tag


        data.append(r_cls)
        data.append(r_au)
        per_teacher_plesso_tags.append(pl_tags)

    end_body = len(data) - 1


    # ========== Excel ==========
    xlsx = OUTPUT_DIR / "ORARIO_TABELLA_GLOBALE.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabella_globale"

    thin = Side(style="thin", color="000000")
    med  = Side(style="medium", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    # usa i fill comuni ovunque
    fill_header, fill_title, alt_fill = excel_get_fills()


    total_cols = 1 + n_time_cols
    fill_header = PatternFill("solid", fgColor=XLSX(COLOR_HEADER_HEX))


    # 1) INTESTAZIONE GLOBALE su righe 1–2 (mergiate e stesso colore delle righe 3–4)
    total_cols = 1 + len(giorni) * len(ore)
    excel_add_global_header(ws, total_cols)
    for r in (1, 2):
        for c in range(1, total_cols+1):
            ws.cell(row=r, column=c).fill = fill_header

    # 2) HEADER Giorni/Ore su righe 3–4
    header_row_top = 3
    header_row_bot = 4

    ws.append(header[0])   # riga 3 (Docente + Giorni "spalmati")
    ws.append(header[1])   # riga 4 ("" + Ore)

    # "Docente" mergiato verticalmente su 3–4
    ws.merge_cells(start_row=header_row_top, start_column=1,
                  end_row=header_row_bot, end_column=1)
    ws.cell(row=header_row_top, column=1, value="Docente")

    # Giorni mergiati orizzontalmente in riga 3, sopra le rispettive ore (riga 4)
    # ATTENZIONE: usare c0 per leggere dalla lista header[0] (indice Python),
    # non start_c che è l'indice di colonna Excel.
    for (c0, c1) in day_bounds:
        start_c = 1 + c0
        end_c   = 1 + c1
        ws.merge_cells(start_row=header_row_top, start_column=start_c,
                      end_row=header_row_top, end_column=end_c)
        ws.cell(row=header_row_top, column=start_c, value=header[0][c0])

    # Stile righe 3–4
    for r in (header_row_top, header_row_bot):
        for c in range(1, total_cols+1):
            ws.cell(row=r, column=c).fill = fill_header
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")

    # Il corpo parte da riga 5
    excel_start_row = header_row_bot + 1
    row_ptr = excel_start_row

    # corpo: due righe per docente
    row_ptr = header_row_bot + 1
    for t_idx, d in enumerate(docenti):
        r_cls = data[start_body + 2*t_idx]
        r_au  = data[start_body + 2*t_idx + 1]
        ws.append(r_cls); ws.append(r_au)

        row_cls = row_ptr
        row_au  = row_ptr + 1
        # merge 'Docente'
        ws.merge_cells(start_row=row_cls, start_column=1, end_row=row_au, end_column=1)

        # evidenzia Succursale (solo se non vuota né 't')
        pl_tags = per_teacher_plesso_tags[t_idx]
        for i, tag in enumerate(pl_tags):
            col_x = 2 + i
            v = tidy(ws.cell(row=row_cls, column=col_x).value)
            if tag == "S" and v and v.lower() != "t":
                ws.cell(row=row_cls, column=col_x).font = Font(bold=True, color=BLUE_HEX)

        # allineamenti
        ws.cell(row=row_cls, column=1).alignment = Alignment(horizontal="left",  vertical="center")
        ws.cell(row=row_au,  column=1).alignment = Alignment(horizontal="left",  vertical="center")
        for c in range(2, total_cols+1):
            ws.cell(row=row_cls, column=c).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            ws.cell(row=row_au,  column=c).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        # [ADD] Zebra striping: colora a blocchi di 2 righe (CLASSI/AULE)
        if t_idx % 2 == 1:  # usa == 0 se vuoi iniziare colorato dalla prima coppia
            for c in range(1, total_cols+1):
                ws.cell(row=row_cls, column=c).fill = alt_fill
                ws.cell(row=row_au,  column=c).fill = alt_fill

        row_ptr += 2

    # bordi + linee fine-giorno
    thin = Side(style="thin", color="000000")
    med  = Side(style="medium", color="000000")
    for r in range(header_row_top, ws.max_row+1):
        for c in range(1, total_cols+1):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for (c0,c1) in day_bounds:
        end_col = 1 + c1
        for r in range(header_row_top, ws.max_row+1):
            cell = ws.cell(row=r, column=end_col)
            cell.border = Border(
                left=cell.border.left, right=med,
                top=cell.border.top,   bottom=cell.border.bottom
            )

    # larghezze colonne semplici
    ws.column_dimensions[get_column_letter(1)].width = 28.0
    for c in range(2, total_cols+1):
        ws.column_dimensions[get_column_letter(c)].width = 14.0

    # === Altezza righe fissa per leggibilità ===
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 17

    wb.save(xlsx)
    return xlsx




# ================================
# EXPORT: Tabella globale per plesso
# ================================
def export_OUTPUT_TABELLA_PLESSO(
    df_all, giorni, ore,
    df_aule=None,
    plesso_focus="Succursale",          # "Centrale" o "Succursale"
    docenti_set=None,                   # se None -> tutti i docenti in df_all
    df_classi=None                      # Tabella_Classi per limitare asterisco/compresenze
):
    import re
    from collections import defaultdict
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import pandas as pd
    from pathlib import Path

    # ---------- helper locali ----------
    def tidy(x):
        if pd.isna(x): return ""
        s = str(x).strip()
        return "" if s.lower() == "nan" else s

    def split_tokens(s: str):
        if not s: return []
        s = str(s).strip()
        for a,b in re.findall(r"\[(.*?)\]|\((.*?)\)", s):
            tok = (a or b or "").strip()
            if tok: s += f" | {tok}"
        return [p.strip() for p in re.split(r"\s*[\|/–—-]\s*", s) if p.strip()]


    def _is_aula(tok: str, known_aule) -> bool:
        t = (tok or "").strip().lower()
        if not t: return False
        if t in known_aule: return True
        return bool(re.match(r"^[A-Za-z]{0,4}\d{2,4}$", t)) or t.startswith("lab") or t.startswith("aula ")


    class_pl_tag = _build_class_plesso_tag(df_classi)

        # --- helper: deduce tag da 'D'/'d' nella cella CLASSI ---
    def _dispo_tag_from_classes_cell(cls_cell: str) -> str | None:
        s = tidy(cls_cell)
        if not s:
            return None
        toks = [t.strip() for t in split_tokens(s) if t.strip()]
        has_D = any(t == "D" for t in toks)   # Centrale
        has_d = any(t == "d" for t in toks)   # Succursale
        if has_D and not has_d:
            return "C"
        if has_d and not has_D:
            return "S"
        return None  # nessun D/d o entrambi (ambiguo)

    # --- helper: classifica la cella CLASSI in 'C'/'S'/None ---
    def _classify_cell_by_classes(cls_cell: str, known_aule) -> str | None:
        # 1) prova con le classi (ignorando token-aula intrusi)
        tags = set()
        for t in split_tokens(tidy(cls_cell)):
            # se hai definito is_aula_token globale, usala; altrimenti fallback locale
            if ('is_aula_token' in globals() and is_aula_token(t, known_aule)) or _is_aula(t, known_aule):
                continue
            cn = _norm_lookup_classe(t)
            if not cn:
                continue
            tg = class_pl_tag.get(cn)
            if tg:
                tags.add(tg)

        if tags == {"C"}:
            return "C"
        if tags == {"S"}:
            return "S"

        # 2) fallback: D/d (disposizione)
        dd = _dispo_tag_from_classes_cell(cls_cell)
        if dd:
            return dd

        return None  # misto/sconosciuto



    # aule note (da df_all + df_aule)
    known_aule = {str(a).strip().lower() for a in df_all["aula"].dropna() if str(a).strip()}
    if df_aule is not None and "Aula" in df_aule.columns:
        known_aule |= {str(a).strip().lower() for a in df_aule["Aula"].dropna() if str(a).strip()}

    # ===== Mappa Aula -> Plesso ('C' o 'S') =====
    room2plesso = {}
    if df_aule is not None and {"Aula","Plesso"} <= set(df_aule.columns):
        for _, r in df_aule.iterrows():
            a = tidy(r.get("Aula", ""))
            p = tidy(r.get("Plesso", "")).lower()
            if not a: continue
            tag = "C" if "centr" in p else ("S" if "succ" in p else None)
            room2plesso[a.lower()] = tag

    # fallback docenti_set
    if docenti_set is None:
        docenti_set = {tidy(d) for d in df_all["docente"].dropna().astype(str) if tidy(d)}

    # set classi valide (limitazione per l'asterisco)
    classes_valid_set = None
    if df_classi is not None and "Classe" in df_classi.columns:
        classes_valid_set = {
            _norm_lookup_classe(c)
            for c in df_classi["Classe"].dropna().astype(str)
            if _norm_lookup_classe(c)
        }

    # ===== Header (2 righe) =====
    header, spans, day_bounds = header_rows_for_day_hour(giorni, ore, first_col_title="Docente")

    # ===== Mapping (docente,giorno,ora) -> CLASSI / AULE + fallback plesso-cell =====
    grp = df_all.groupby(["docente","giorno","ora"], dropna=False, observed=False)
    classes_map, aule_map = defaultdict(list), defaultdict(list)
    cell_plessi_map = {}  # (d,g,o) -> set di plessi sorgente ("Centrale"/"Succursale")

    for (d,g,o), sub in grp:
        key = (str(d), str(g), str(o))
        for raw in sub["classe"].dropna().tolist():
            for tok in split_tokens(raw):
                (aule_map if _is_aula(tok, known_aule) else classes_map)[key].append(tok)
        for raw in sub["aula"].dropna().tolist():
            for tok in split_tokens(raw):
                aule_map[key].append(tok)
        # plessi sorgente per fallback (se l'aula non aiuta)
        if "plesso" in sub.columns:
            cell_plessi_map[key] = {tidy(p) for p in sub["plesso"].astype(str).unique() if tidy(p)}
        else:
            cell_plessi_map[key] = set()

    classes_map = {k: " | ".join(sorted(set(v))) for k,v in classes_map.items()}
    aule_map    = {k: " | ".join(sorted(set(v))) for k,v in aule_map.items()}

    # ===== Docenti da includere: presenti nel plesso_focus (almeno un'ora non vuota) =====
    mask_focus = df_all["plesso"].astype(str).str.lower() == plesso_focus.lower()
    mask_nonvuoto = (df_all["classe"].astype(str).str.strip() != "") | (df_all["aula"].astype(str).str.strip() != "")
    docenti_focus = sorted(
        {d.strip() for d in df_all.loc[mask_focus & mask_nonvuoto, "docente"].astype(str) if d and d.strip()},
        key=lambda s: s.lower()
    )


    # --------- Classifica cella -> 'C' / 'S' / None (usa D/d, poi aula, poi fallback sorgente) ---------
    def classify_plesso_cell(key, aule_cell: str, cls_cell: str) -> str | None:
        # 1) priorità a D/d nella cella CLASSI
        dd = _dispo_tag_from_classes_cell(cls_cell)
        if dd in {"C", "S"}:
            return dd

        # 2) deduci dal plesso delle AULE (room2plesso)
        s = tidy(aule_cell)
        tags = set()
        if s:
            for t in [t.strip() for t in s.split("|") if t.strip()]:
                tag = room2plesso.get(t.lower())
                if tag in {"C","S"}:
                    tags.add(tag)
        if tags == {"C"}: return "C"
        if tags == {"S"}: return "S"

        # 3) fallback: plessi sorgente univoci
        pl_full_set = cell_plessi_map.get(key, set())
        if len(pl_full_set) == 1:
            only = next(iter(pl_full_set)).lower()
            if "centr" in only: return "C"
            if "succ"  in only: return "S"

        return None


    # --------- PRECALCOLO COMPRESENZE (per asterisco) ---------
    comp_map = defaultdict(set)  # chiave: (giorno, ora, classe_norm, aula_norm)

    def _is_aula_token(tok: str) -> bool:
        return _is_aula(tok, known_aule)


    for _, r in df_all.iterrows():
        g  = tidy(r.get("giorno",""))
        o  = tidy(r.get("ora",""))
        d  = tidy(r.get("docente",""))
        if not (g and o and d): continue
        # escludi sostegno
        if str(r.get("is_sostegno","")).strip().lower() in {"true","1"}:
            continue

        class_norms = []
        for tok in split_tokens(tidy(r.get("classe",""))):
            if _is_aula_token(tok): continue
            cn = _norm_lookup_classe(tok)
            if not cn: continue
            if classes_valid_set is not None and cn not in classes_valid_set: continue
            class_norms.append(cn)
        if not class_norms: continue

        aule_tokens = set()
        for tok in split_tokens(tidy(r.get("aula",""))):
            if _is_aula_token(tok): aule_tokens.add(tok.strip().lower())
        for tok in split_tokens(tidy(r.get("classe",""))):
            if _is_aula_token(tok): aule_tokens.add(tok.strip().lower())
        if not aule_tokens: continue

        for cn in class_norms:
            for a_norm in aule_tokens:
                comp_map[(str(g), str(o), cn, a_norm)].add(d)

    comp_keys = {k for k, ds in comp_map.items() if len(ds) >= 2}



    # ====== SOLO EXCEL ======
    xlsx = OUTPUT_DIR / f"ORARIO_TABELLA_{plesso_focus.upper()}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Tabella_{plesso_focus.lower()}"

    thin   = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # FILL header/zebra con fallback robusti
    if "XLSX" in globals() and "COLOR_HEADER_HEX" in globals():
        fill_header = PatternFill("solid", fgColor=XLSX(COLOR_HEADER_HEX))
    else:
        fill_header = PatternFill("solid", fgColor="D9E3F0")
    if "XLSX" in globals() and "COLOR_ZEBRA_ALT_HEX" in globals():
        alt_fill = PatternFill("solid", fgColor=XLSX(COLOR_ZEBRA_ALT_HEX))
    else:
        alt_fill = PatternFill("solid", fgColor="EEF4FF")  # azzurrino di default

    # colonne totali: 1 ("Docente") + giorni*ore
    n_hours = len(ore)
    ncols_total = 1 + len(giorni) * n_hours

    # === 1) INTESTAZIONE GLOBALE (righe 1–2) ===
    header_txt = (HEADER_TEXT or "") if "HEADER_TEXT" in globals() else ""
    ws.append([header_txt]); ws.append([""])
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=ncols_total)
    cell_title = ws.cell(row=1, column=1)
    cell_title.font = Font(bold=True, italic=True)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    for r in (1, 2):
        for c in range(1, ncols_total+1):
            ws.cell(row=r, column=c).fill = fill_header

    # === 2) HEADER Giorni/Ore (righe 3–4) ===
    header_top, header_bot = header
    header_row_top = 3
    header_row_bot = 4
    ws.append(header_top)  # riga 3
    ws.append(header_bot)  # riga 4

    ws.merge_cells(start_row=header_row_top, start_column=1, end_row=header_row_bot, end_column=1)  # "Docente"

    # Giorni mergiati orizzontalmente (offset +1 per la colonna Docente) – indice corretto
    for (c0, c1) in day_bounds:
        start_c = 1 + c0
        end_c   = 1 + c1
        ws.merge_cells(start_row=header_row_top, start_column=start_c, end_row=header_row_top, end_column=end_c)
        ws.cell(row=header_row_top, column=start_c, value=header_top[c0])

    for r in (header_row_top, header_row_bot):
        for c in range(1, ncols_total+1):
            ws.cell(row=r, column=c).fill = fill_header
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")

    # === Corpo ===
    cur_row = header_row_bot + 1
    focus_tag = "C" if plesso_focus.lower().startswith("centr") else "S"

    ws.column_dimensions[get_column_letter(1)].width = 26.0

    for t_idx, d in enumerate(docenti_focus):
        r_cls = [d]
        r_au  = [""]

        plesso_tags = []
        for g in giorni:
            for o in ore:
                k = (str(d), str(g), str(o))
                cls_txt = classes_map.get(k, "")
                aul_txt = aule_map.get(k, "")

                # asterischi compresenza (usa aula -> plesso; se vuoto, fallback ai plessi sorgente)
                plessi_cella_full = set()
                for t in split_tokens(aul_txt):
                    tag = room2plesso.get(t.strip().lower())
                    if tag == "C": plessi_cella_full.add("Centrale")
                    elif tag == "S": plessi_cella_full.add("Succursale")
                if not plessi_cella_full:
                    plessi_cella_full = cell_plessi_map.get(k, set())

                new_cls_tokens = []
                for tkn in split_tokens(cls_txt):
                    if _is_aula(tkn, known_aule):
                        new_cls_tokens.append(tkn); continue
                    cn = _norm_lookup_classe(tkn)
                    if not cn or (classes_valid_set is not None and cn not in classes_valid_set):
                        new_cls_tokens.append(tkn); continue
                    # aule effettive nella cella corrente
                    aule_cell = set()
                    for tok_a in split_tokens(aul_txt):
                        if _is_aula(tok_a, known_aule): aule_cell.add(tok_a.strip().lower())
                    for tok_a in split_tokens(cls_txt):
                        if _is_aula(tok_a, known_aule): aule_cell.add(tok_a.strip().lower())


                    starred = any((str(g), str(o), cn, a_norm) in comp_keys for a_norm in aule_cell)
                    new_cls_tokens.append(re.sub(r"\*+$", "", tkn) + ("*" if starred else ""))


                cls_txt = " | ".join(new_cls_tokens)

                plesso_tags.append(_classify_cell_by_classes(cls_txt, known_aule))
                r_cls.append(cls_txt)
                r_au.append(aul_txt)

        # logica 't' (marca cella vuota prima del cambio plesso nello stesso giorno)
        for di, g in enumerate(giorni):
            base = di * n_hours
            last_tag = None
            for h in range(0, n_hours):
                cur_tag = plesso_tags[base + h]
                if cur_tag in {"C","S"}:
                    if last_tag is None:
                        last_tag = cur_tag
                    elif cur_tag != last_tag:
                        j = h
                        if j - 1 >= 0:
                            target_idx = 1 + base + (j - 1)
                            if tidy(r_cls[target_idx]) == "": r_cls[target_idx] = "t"
                        last_tag = cur_tag
                    else:
                        last_tag = cur_tag

        # scrivi riga CLASSI + AULE
        ws.append(r_cls)
        ws.append(r_au)

        # zebra striping per coppie (CLASSI/AULE)
        if t_idx % 2 == 1:   # colora la 2a, 4a, ...
            for c in range(1, ncols_total+1):
                ws.cell(row=cur_row,   column=c).fill = alt_fill     # riga CLASSI
                ws.cell(row=cur_row+1, column=c).fill = alt_fill     # riga AULE

        # merge prima colonna sulle due righe
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row+1, end_column=1)
        ws.cell(row=cur_row, column=1).alignment = Alignment(vertical="center")

        # wrap/align sulle colonne tempo
        for rr in (cur_row, cur_row+1):
            for cc in range(2, ncols_total+1):
                ws.cell(row=rr, column=cc).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        # evidenzia celle CLASSI dell'altro plesso (usa classify con fallback)
        for di, g in enumerate(giorni):
            for hi, o in enumerate(ore):
                col_idx = 2 + di*n_hours + hi
                aul_txt = ws.cell(row=cur_row+1, column=col_idx).value or ""
                k_cur = (str(d), str(g), str(o))
                cls_val = ws.cell(row=cur_row, column=col_idx).value or ""
                tag = classify_plesso_cell(k_cur, aul_txt, cls_val)
                is_other = (tag is not None) and (tag != focus_tag)
                if is_other and cls_val and cls_val != "t":
                    ws.cell(row=cur_row, column=col_idx).font = Font(bold=True, color="1F4E79")

        cur_row += 2

    end_row = cur_row - 1

    # === Altezza righe fissa per leggibilità ===
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 17

    # Bordi + linee spesse
    for r in range(1, end_row+1):
        for c in range(1, ncols_total+1):
            ws.cell(row=r, column=c).border = border_all
    # linea più spessa dopo la colonna 'Docente'
    for r in range(1, end_row+1):
        cell = ws.cell(row=r, column=1)
        cell.border = Border(
            left=cell.border.left, right=medium,
            top=cell.border.top, bottom=cell.border.bottom
        )
    # linee verticali spesse a fine giorno
    for i_g, g in enumerate(giorni, start=1):
        c_end = 1 + i_g*len(ore)
        for r in range(1, end_row+1):
            cell = ws.cell(row=r, column=c_end)
            cell.border = Border(
                left=cell.border.left, right=medium,
                top=cell.border.top, bottom=cell.border.bottom
            )

    # larghezze colonne (stima semplice)
    max_len = 0
    for r in range(3, end_row+1):
        for c in range(2, ncols_total+1):
            v = ws.cell(row=r, column=c).value
            if v: max_len = max(max_len, len(str(v)))
    time_w = max(10.0, min(28.0, max_len * 0.6 + 2))
    for c in range(2, ncols_total+1):
        ws.column_dimensions[get_column_letter(c)].width = time_w

    wb.save(xlsx)
    print("Creato XLSX:", xlsx)
    return xlsx




# ================================
# EXPORT Aule compatto
# ================================

from pathlib import Path

def export_OUTPUT_AULE_COMPATTO(
    df_plesso,
    titolo="OUTPUT_AULE_COMPATTO",
    base_font=8.5,
    # === EXCEL: controlli larghezze ===
    xlsx_first_col_width=18.0,       # caratteri
    xlsx_time_col_width=None,        # float -> tutte uguali; None => day/slot/auto
    xlsx_day_col_widths=None,        # dict: {"Lunedì": 15.0, ...}
    xlsx_slot_col_widths=None,       # dict: {("Lunedì", 1): 18.0, ...}
    xlsx_time_col_min=10.0,
    xlsx_time_col_max=45.0,
    *,
    df_aule=None,                      # <--- AGGIUNTO
):
    """
    Righe = Aule; Colonne = Giorno/Ora.
    Celle = Docente (riga 1) + Classe (riga 2) nella stessa cella, con a capo.
    Excel: titolo = HEADER_TEXT su prime 2 righe mergiate; prime 4 righe azzurre;
           zebra striping sul corpo; altezza righe stimata sui contenuti.
    """
    from collections import defaultdict
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import pandas as pd

    def tidy(x):
        if pd.isna(x): return ""
        s = str(x).strip()
        return "" if s.lower()=="nan" else s

    # ------------------- dati base -------------------
    giorni_all = list(df_plesso["giorno"].cat.categories)
    ore_all    = list(df_plesso["ora"].cat.categories)

    aule = sorted({a for a in df_plesso["aula"].unique() if tidy(a)}) or ["(nessuna aula rilevata)"]
    # lookup capienza per aula (case-insensitive) da Tabella_Aule
    capienza_map_ci = {}
    if df_aule is not None and {"Aula", "Capienza"} <= set(df_aule.columns):
        for _, r in df_aule.iterrows():
            a_name = tidy(r.get("Aula", ""))
            if a_name:
                capienza_map_ci[a_name.strip().lower()] = r.get("Capienza")

    # mapping (aula,giorno,ora) -> docenti / classi
    mapping_doc  = defaultdict(set)
    mapping_cls  = defaultdict(set)
    for _, r in df_plesso.iterrows():
        aula = tidy(r["aula"])
        if not aula:
            continue
        key = (aula, str(r["giorno"]), str(r["ora"]))
        d   = tidy(r["docente"])
        c   = tidy(r["classe"])
        if d: mapping_doc[key].add(d)
        if c: mapping_cls[key].add(c)

    # ------------------- EXCEL -------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aule"

    giorni_subset = list(giorni_all)   # in Excel: tutti i giorni
    ore_subset    = list(ore_all)

    total_time_cols = len(giorni_subset) * len(ore_subset)
    total_cols = 1 + total_time_cols   # col 1 = Aule



    # === Header a due righe (giorni/ore) ===
    excel_add_global_header(ws, total_cols)

    header_row_top = 3
    header_row_bot = 4
    ws.merge_cells(start_row=header_row_top, start_column=1, end_row=header_row_bot, end_column=1)
    ws.cell(row=header_row_top, column=1, value="Aule").alignment = Alignment(horizontal="center", vertical="center")

    col = 2
    for g in giorni_subset:
        start = col
        for o in ore_subset:
            ws.cell(row=header_row_bot, column=col, value=str(o)).alignment = Alignment(horizontal="center", vertical="center")
            col += 1
        end = col - 1
        ws.merge_cells(start_row=header_row_top, start_column=start, end_row=header_row_top, end_column=end)
        ws.cell(row=header_row_top, column=start, value=str(g)).alignment = Alignment(horizontal="center", vertical="center")


    # === Colora prime 4 righe in azzurro header ===
    fill_header = PatternFill("solid", fgColor=XLSX(COLOR_HEADER_HEX))
    for r in range(1, 5):
        for c in range(1, total_cols+1):
            ws.cell(row=r, column=c).fill = fill_header

    # === Corpo (una riga per aula; celle = "DOCENTE\\nclasse") ===
    start_body = header_row_bot + 1
    row = start_body
    for aula in aule:
        cap = capienza_map_ci.get(aula.strip().lower())
        label = f"{aula} ({cap})" if cap not in (None, "", "nan") else aula
        ws.cell(row=row, column=1, value=label).alignment = Alignment(horizontal="center", vertical="center")

        col = 2
        for g in giorni_subset:
            for o in ore_subset:
                key = (aula, str(g), str(o))
                doc_txt = " | ".join(sorted(mapping_doc.get(key, set())))
                cls_txt = " | ".join(sorted({x for x in mapping_cls.get(key, set()) if x}))
                if doc_txt and cls_txt:
                    val = f"{doc_txt}\n{cls_txt}"
                else:
                    val = doc_txt or cls_txt
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                col += 1
        row += 1
    end_body = row - 1

    # === Bordi e linee fine-giorno ===
    thin   = Side(style="thin",   color=XLSX(COLOR_GRID_HEX))
    medium = Side(style="medium", color=XLSX(COLOR_GRID_HEX))
    for r in range(3, end_body+1):
        for c in range(1, total_cols+1):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # linee verticali spesse a fine giorno
    for i_g, g in enumerate(giorni_subset, start=1):
        c_end = 1 + i_g*len(ore_subset)  # 1 = col Aule
        for r in range(3, end_body+1):
            cell = ws.cell(row=r, column=c_end)
            cell.border = Border(
                left=cell.border.left, right=medium,
                top=cell.border.top,   bottom=cell.border.bottom
            )

    # bordo verticale spesso dopo la prima colonna
    for r in range(3, end_body + 1):
        cell = ws.cell(row=r, column=1)
        cell.border = Border(
            left   = cell.border.left,
            right  = medium,
            top    = cell.border.top,
            bottom = cell.border.bottom,
        )

    # bordo orizzontale spesso dopo la 4ª riga
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=4, column=c)
        cell.border = Border(
            left   = cell.border.left,
            right  = cell.border.right,
            top    = cell.border.top,
            bottom = medium,
        )

    # === Larghezze colonne Excel ===
    ws.column_dimensions[get_column_letter(1)].width = float(xlsx_first_col_width)
    # mappa colonna -> (giorno, ora)
    idx_to_slot = {}
    c = 2
    for g in giorni_subset:
        for o in ore_subset:
            idx_to_slot[c] = (str(g), str(o))
            c += 1

    if xlsx_slot_col_widths:  # priorità 1
        for col_idx, (g, o) in idx_to_slot.items():
            for k in ((g, o), (g, int(o) if o.isdigit() else o)):
                if k in xlsx_slot_col_widths:
                    ws.column_dimensions[get_column_letter(col_idx)].width = float(xlsx_slot_col_widths[k])
                    break
    if xlsx_day_col_widths:   # priorità 2
        for col_idx, (g, _) in idx_to_slot.items():
            if g in xlsx_day_col_widths:
                ws.column_dimensions[get_column_letter(col_idx)].width = float(xlsx_day_col_widths[g])
    if xlsx_time_col_width is not None:  # priorità 3
        for col_idx in idx_to_slot.keys():
            ws.column_dimensions[get_column_letter(col_idx)].width = float(xlsx_time_col_width)
    if xlsx_time_col_width is None and not xlsx_day_col_widths and not xlsx_slot_col_widths:
        # auto: stima dalla lunghezza massima dei contenuti
        max_chars = 0
        for r in range(start_body, end_body+1):
            for col_idx in idx_to_slot.keys():
                v = ws.cell(row=r, column=col_idx).value
                if v:
                    max_chars = max(max_chars, len(str(v)))
        auto_w = max(xlsx_time_col_min, min(xlsx_time_col_max, max_chars*0.6 + 2))
        for col_idx in idx_to_slot.keys():
            ws.column_dimensions[get_column_letter(col_idx)].width = float(auto_w)

    # === Zebra striping sul corpo ===
    alt_fill = PatternFill("solid", fgColor=XLSX(COLOR_ZEBRA_ALT_HEX))
    for r in range(start_body, end_body+1):
        if (r - start_body) % 2 == 1:
            for c in range(1, total_cols+1):
                ws.cell(row=r, column=c).fill = alt_fill

    # === Altezza righe: stima in base ai contenuti ===
    def _estimate_row_height(r_idx: int) -> float:
        max_lines = 1
        for c_idx in range(1, total_cols+1):
            v = ws.cell(row=r_idx, column=c_idx).value
            if v is None:
                continue
            txt = str(v)
            # base: righe esplicite
            lines = txt.count("\n") + 1
            # stima wrap (chars per riga ≈ col width)
            col_w = ws.column_dimensions[get_column_letter(c_idx)].width or 10
            seg_lines = 0
            for seg in txt.split("\n"):
                seg_lines += max(1, int((len(seg) + col_w - 1) // col_w))
            lines = max(lines, seg_lines)
            max_lines = max(max_lines, lines)
        # coeff ~14.5pt per linea con font 8.5 + un piccolo margine
        return max(18.0, max_lines * 14.5)

    for r in range(start_body, end_body+1):
        ws.row_dimensions[r].height = _estimate_row_height(r)

    # salva Excel (accetta anche 'titolo' passato come .pdf e lo converte a .xlsx)
    t = str(titolo)
    if t.lower().endswith(".pdf"):
        t = t[:-4]
    if not t.lower().endswith(".xlsx"):
        t = t + ".xlsx"
    xlsx_path = OUTPUT_DIR / t
    wb.save(xlsx_path)

    print("Creato XLSX:", xlsx_path)
    return xlsx_path





# ================================
# EXPORT 5-6: Classi compatto (righe=classi, celle=aule)
# ================================
from pathlib import Path

# ================================
# EXPORT 5-6: Classi compatto (righe=classi, celle=aule) — SOLO EXCEL
# ================================
def export_OUTPUT_CLASSI_COMPATTO(
    df_plesso,
    titolo="OUTPUT_CLASSI_COMPATTO",
    # === EXCEL: controlli larghezze ===
    xlsx_first_col_width=16.0,       # caratteri
    xlsx_time_col_width=None,        # float -> tutte uguali; None => auto
    xlsx_day_col_widths=None,        # dict: {"Lunedì": 15.0, ...}
    xlsx_slot_col_widths=None,       # dict: {("Lunedì", 1): 18.0, ...}
    xlsx_time_col_min=10.0,
    xlsx_time_col_max=45.0,
):
    import re
    from collections import defaultdict
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import pandas as pd

    # ---------- helper ----------
    def tidy(x):
        if pd.isna(x): return ""
        s = str(x).strip()
        return "" if s.lower()=="nan" else s

    def split_tokens(s: str):
        """Split robusto su | / – — - ; porta in chiaro [..] e (..)."""
        if not s: return []
        s = str(s).strip()
        for a,b in re.findall(r"\[(.*?)\]|\((.*?)\)", s):
            tok = (a or b or "").strip()
            if tok: s += f" | {tok}"
        return [p.strip() for p in re.split(r"\s*[\|/–—-]\s*", s) if p.strip()]

    def norm_classe(tok: str) -> str:
        """Rimuove eventuali asterischi finali e spazi."""
        return re.sub(r"\*+$", "", (tok or "").strip())

    # ---------- Dati base ----------
    giorni_all = list(df_plesso["giorno"].cat.categories)
    ore_all    = list(df_plesso["ora"].cat.categories)

    # Mapping aggregati su classe NORMALIZZATA (senza asterisco)
    mapping_doc  = defaultdict(set)  # (cls_norm, giorno, ora) -> {docenti}
    mapping_aule = defaultdict(set)  # (cls_norm, giorno, ora) -> {aule}
    classi_set   = set()

    for _, r in df_plesso.iterrows():
        raw_cls = tidy(r["classe"])
        if not raw_cls:
            continue
        g = str(r["giorno"]); o = str(r["ora"])
        d = tidy(r["docente"]); a = tidy(r["aula"])
        for tok in split_tokens(raw_cls):
            cls_norm = norm_classe(tok)
            if not cls_norm:
                continue
            classi_set.add(cls_norm)
            key = (cls_norm, g, o)
            if d: mapping_doc[key].add(d)
            if a: mapping_aule[key].add(a)

    # elenco classi (senza asterischi), ordinato
    classi = sorted(classi_set) or ["(nessuna classe rilevata)"]

    # ====== EXCEL ======
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classi"

    giorni_subset = list(giorni_all)
    ore_subset    = list(ore_all)

    total_time_cols = len(giorni_subset) * len(ore_subset)
    total_cols = 1 + total_time_cols   # col 1 = Classi

    # === Titolo su prime 2 righe mergiate ===
    header_text = (HEADER_TEXT or "") if "HEADER_TEXT" in globals() else ""
    ws.append([header_text])
    ws.append([""])
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=total_cols)
    tit = ws.cell(row=1, column=1)
    tit.alignment = Alignment(horizontal="center", vertical="center")
    tit.font = Font(bold=True, italic=True)

    # === Header a due righe ===
    header_row_top = 3
    header_row_bot = 4
    ws.merge_cells(start_row=header_row_top, start_column=1, end_row=header_row_bot, end_column=1)
    ws.cell(row=header_row_top, column=1, value="Classi").alignment = Alignment(horizontal="center", vertical="center")

    col = 2
    for g in giorni_subset:
        start = col
        for o in ore_subset:
            ws.cell(row=header_row_bot, column=col, value=str(o)).alignment = Alignment(horizontal="center", vertical="center")
            col += 1
        end = col - 1
        ws.merge_cells(start_row=header_row_top, start_column=start, end_row=header_row_top, end_column=end)
        ws.cell(row=header_row_top, column=start, value=str(g)).alignment = Alignment(horizontal="center", vertical="center")

    # === Colora prime 4 righe in azzurro header ===
    fill_header = PatternFill("solid", fgColor=XLSX(COLOR_HEADER_HEX))
    for r in range(1, 5):
        for c in range(1, total_cols+1):
            ws.cell(row=r, column=c).fill = fill_header

    # === Corpo (una riga per classe; celle = "DOCENTE\nAula") ===
    start_body = header_row_bot + 1
    row = start_body
    for cls in classi:
        ws.cell(row=row, column=1, value=cls).alignment = Alignment(horizontal="center", vertical="center")
        col = 2
        for g in giorni_subset:
            for o in ore_subset:
                key = (cls, str(g), str(o))
                doc_txt  = " | ".join(sorted(mapping_doc.get(key, set())))
                aula_txt = " | ".join(sorted(mapping_aule.get(key, set())))
                if doc_txt and aula_txt:
                    val = f"{doc_txt}\n{aula_txt}"
                else:
                    val = doc_txt or aula_txt
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                col += 1
        row += 1
    end_body = row - 1

    # === Bordi e linee fine-giorno ===
    thin   = Side(style="thin",   color=XLSX(COLOR_GRID_HEX))
    medium = Side(style="medium", color=XLSX(COLOR_GRID_HEX))
    for r in range(3, end_body+1):
        for c in range(1, total_cols+1):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # separatori verticali a fine giorno
    for i_g, _ in enumerate(giorni_subset, start=1):
        c_end = 1 + i_g*len(ore_subset)  # 1 = col Classi
        for r in range(3, end_body+1):
            cell = ws.cell(row=r, column=c_end)
            cell.border = Border(
                left=cell.border.left, right=medium,
                top=cell.border.top,   bottom=cell.border.bottom
            )
    # bordo verticale spesso dopo la prima colonna
    for r in range(3, end_body + 1):
        cell = ws.cell(row=r, column=1)
        cell.border = Border(
            left   = cell.border.left,
            right  = medium,
            top    = cell.border.top,
            bottom = cell.border.bottom,
        )
    # bordo orizzontale spesso dopo la 4ª riga
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=4, column=c)
        cell.border = Border(
            left   = cell.border.left,
            right  = cell.border.right,
            top    = cell.border.top,
            bottom = medium,
        )

    # === Larghezze colonne Excel ===
    ws.column_dimensions[get_column_letter(1)].width = float(xlsx_first_col_width)
    # mappa colonna -> (giorno, ora)
    idx_to_slot = {}
    c = 2
    for g in giorni_subset:
        for o in ore_subset:
            idx_to_slot[c] = (str(g), str(o))
            c += 1

    if xlsx_slot_col_widths:  # priorità 1
        for col_idx, (g, o) in idx_to_slot.items():
            for k in ((g, o), (g, int(o) if o.isdigit() else o)):
                if k in xlsx_slot_col_widths:
                    ws.column_dimensions[get_column_letter(col_idx)].width = float(xlsx_slot_col_widths[k])
                    break
    if xlsx_day_col_widths:   # priorità 2
        for col_idx, (g, _) in idx_to_slot.items():
            if g in xlsx_day_col_widths:
                ws.column_dimensions[get_column_letter(col_idx)].width = float(xlsx_day_col_widths[g])
    if xlsx_time_col_width is not None:  # priorità 3
        for col_idx in idx_to_slot.keys():
            ws.column_dimensions[get_column_letter(col_idx)].width = float(xlsx_time_col_width)
    if xlsx_time_col_width is None and not xlsx_day_col_widths and not xlsx_slot_col_widths:
        # auto: stima dalla lunghezza massima dei contenuti
        max_chars = 0
        for r in range(start_body, end_body+1):
            for col_idx in idx_to_slot.keys():
                v = ws.cell(row=r, column=col_idx).value
                if v:
                    max_chars = max(max_chars, len(str(v)))
        auto_w = max(xlsx_time_col_min, min(xlsx_time_col_max, max_chars*0.6 + 2))
        for col_idx in idx_to_slot.keys():
            ws.column_dimensions[get_column_letter(col_idx)].width = float(auto_w)

    # === Zebra striping sul corpo ===
    alt_fill = PatternFill("solid", fgColor=XLSX(COLOR_ZEBRA_ALT_HEX))
    for r in range(start_body, end_body+1):
        if (r - start_body) % 2 == 1:
            for c in range(1, total_cols+1):
                ws.cell(row=r, column=c).fill = alt_fill

    # === Altezza righe: stima in base ai contenuti ===
    def _estimate_row_height(r_idx: int) -> float:
        max_lines = 1
        for c_idx in range(1, total_cols+1):
            v = ws.cell(row=r_idx, column=c_idx).value
            if v is None:
                continue
            txt = str(v)
            # base: righe esplicite
            lines = txt.count("\n") + 1
            # stima wrap (chars per riga ≈ col width)
            col_w = ws.column_dimensions[get_column_letter(c_idx)].width or 10
            seg_lines = 0
            for seg in txt.split("\n"):
                seg_lines += max(1, int((len(seg) + col_w - 1) // col_w))
            lines = max(lines, seg_lines)
            max_lines = max(max_lines, lines)
        # ~14.5pt per linea con font 8.5 + margine
        return max(18.0, max_lines * 14.5)

    for r in range(start_body, end_body+1):
        ws.row_dimensions[r].height = _estimate_row_height(r)

    # salva Excel (accetta anche 'titolo' passato come .pdf e lo converte a .xlsx)
    t = str(titolo)
    if t.lower().endswith(".pdf"):
        t = t[:-4]
    if not t.lower().endswith(".xlsx"):
        t = t + ".xlsx"
    xlsx_path = OUTPUT_DIR / t
    wb.save(xlsx_path)

    print("Creato XLSX:", xlsx_path)
    return xlsx_path

# ================================
# Aule libere
# ================================

def export_OUTPUT_AULE_LIBERE(
    df_plesso,
    df_aule,
    plesso_label: str,
    nome_file_xlsx: str,
    # larghezze colonne
    xlsx_giorno_col_width=11.0,
    xlsx_ora_col_width=6.0,
    xlsx_aule_col_width=16.0,
):
    """
    Crea un Excel 'aule libere' con:
    - Col A (Giorno) mergiata su tutte le (ore * 2) righe del giorno
    - Col B (Ora) mergiata su 2 righe per ciascuna ora (seconda riga vuota)
    - Riga 3: 'Giorno' | 'Ora' | 'Aula 1' ... 'Aula N'
    - Riga 3 e riga separatrice tra i giorni con bordo inferiore 'medium'
    """
    import re
    from pathlib import Path
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # ---- util locali ----
    def tidy(x):
        import pandas as pd
        if pd.isna(x): return ""
        s = str(x).strip()
        return "" if s.lower() == "nan" else s

    def split_tokens(s: str):
        if not s: return []
        s = str(s).strip()
        # porta in chiaro [..] e (..)
        for a,b in re.findall(r"\[(.*?)\]|\((.*?)\)", s):
            tok = (a or b or "").strip()
            if tok: s += f" | {tok}"
        return [p.strip() for p in re.split(r"\s*[\|/–—-]\s*", s) if p.strip()]

    # ---- asse giorni/ore ----
    if "giorno" not in df_plesso.columns or "ora" not in df_plesso.columns or "aula" not in df_plesso.columns:
        raise ValueError("df_plesso deve contenere le colonne 'giorno', 'ora', 'aula'.")

    giorni_order = list(df_plesso["giorno"].cat.categories) if str(df_plesso["giorno"].dtype).startswith("category") \
                   else [g for g in sorted(df_plesso["giorno"].astype(str).unique(), key=str)]
    ore_order    = list(df_plesso["ora"].cat.categories)    if str(df_plesso["ora"].dtype).startswith("category") \
                   else sorted(df_plesso["ora"].astype(str).unique(), key=lambda x: (len(x), x))

    # ---- aule del plesso ----
    if df_aule is None or not {"Aula","Plesso"} <= set(df_aule.columns):
        raise ValueError("df_aule deve avere colonne 'Aula' e 'Plesso'.")

    aule_plesso = [tidy(r.get("Aula","")) for _, r in df_aule.iterrows()
                   if tidy(r.get("Plesso","")).lower() == plesso_label.lower() and tidy(r.get("Aula",""))]
    aule_plesso = sorted(dict.fromkeys(aule_plesso))  # uniche, ordinate
    if not aule_plesso:
        raise ValueError(f"Nessuna aula trovata per il plesso '{plesso_label}' in df_aule.")

    # --- lookup capienza per aula del plesso corrente ---
    def _safe_int(x):
        s = str(x).strip()
        if not s or s.lower() == "nan":
            return None
        import re
        m = re.search(r"\d+", s)
        return int(m.group(0)) if m else None

    capienza_map_ci = {}
    for _, r in df_aule.iterrows():
        if tidy(r.get("Plesso","")).strip().lower() == plesso_label.strip().lower():
            a_name = tidy(r.get("Aula",""))
            if a_name:
                capienza_map_ci[a_name.strip().lower()] = _safe_int(r.get("Capienza"))

    # ---- mappa (giorno, ora) -> aule occupate ----
    used_by_slot = {}
    for g in giorni_order:
        for o in ore_order:
            sub = df_plesso[(df_plesso["giorno"].astype(str)==str(g)) &
                            (df_plesso["ora"].astype(str)==str(o))]
            occ = set()
            for raw in sub["aula"].astype(str).tolist():
                for t in split_tokens(raw):
                    if t: occ.add(t)
            used_by_slot[(str(g), str(o))] = occ

    # ---- numero colonne Aula (max aule libere osservate) ----
    max_free = 0
    for g in giorni_order:
        for o in ore_order:
            free = [a for a in aule_plesso if a not in used_by_slot[(str(g), str(o))]]
            if len(free) > max_free: max_free = len(free)
    max_free = max(1, max_free)

    # ---- workbook ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Aule_libere_{plesso_label}"

    total_cols = 2 + max_free  # Giorno, Ora, Aula 1..N

    # riga 1-2: intestazione globale
    excel_add_global_header(ws, total_cols)

    # riga 3: intestazioni
    header_row = 3
    titles = ["Giorno", "Ora"] + [f"Aula {i}" for i in range(1, max_free+1)]
    ws.append(titles)
    for c in range(1, total_cols+1):
        ws.cell(row=header_row, column=c).alignment = Alignment(horizontal="center", vertical="center")

    # colori/bordi
    grid_hex = XLSX(COLOR_GRID_HEX) if "XLSX" in globals() and "COLOR_GRID_HEX" in globals() else "000000"
    fill_header = PatternFill("solid", fgColor=XLSX(COLOR_HEADER_HEX)) if "XLSX" in globals() and "COLOR_HEADER_HEX" in globals() else None
    thin   = Side(style="thin",   color=grid_hex)
    medium = Side(style="medium", color=grid_hex)

    # colora riga 1-3 come header
    for r in (1, 2, 3):
        for c in range(1, total_cols+1):
            cell = ws.cell(row=r, column=c)
            if fill_header: cell.fill = fill_header

    # bordo inferiore doppio sulla riga 3
    for c in range(1, total_cols+1):
        cell = ws.cell(row=3, column=c)
        cell.border = Border(left=thin, right=thin, top=cell.border.top if cell.border else thin, bottom=medium)

    # ---- corpo: per ogni giorno/ora 2 righe ----
    current_row = 3
    for g in giorni_order:
        day_start_row = current_row + 1  # prima riga del blocco (dopo header)

        for o in ore_order:
            occ = used_by_slot[(str(g), str(o))]
            free_list = [a for a in aule_plesso if a not in occ]
            padded = free_list + [""]*(max_free - len(free_list))

            # riga dati
            current_row += 1
            ws.cell(row=current_row, column=1, value=str(g))
            try:
                start_time = ORE_MAP.get(int(o), str(o)) if isinstance(o, (int, float, str)) else str(o)
            except Exception:
                start_time = str(o)
            ws.cell(row=current_row, column=2, value=str(start_time))

            # >>>>>>>>>>>>>>>>>> QUI LA MODIFICA PER MOSTRARE LA CAPIENZA <<<<<<<<<<<<<<<<<<
            for i, val in enumerate(padded, start=3):
                aula_clean = tidy(val)
                if aula_clean:
                    cap = capienza_map_ci.get(aula_clean.strip().lower())
                    label = f"{aula_clean} ({cap})" if cap is not None else aula_clean
                else:
                    label = ""
                ws.cell(row=current_row, column=i, value=label)
            # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

            # riga vuota sotto
            current_row += 1
            for c in range(1, total_cols+1):
                ws.cell(row=current_row, column=c, value="")

            # merge 'Ora' sulle 2 righe
            ws.merge_cells(start_row=current_row-1, start_column=2, end_row=current_row, end_column=2)
            ws.cell(row=current_row-1, column=2).alignment = Alignment(horizontal="center", vertical="center")

            # bordi sottili per le due righe dell'ora
            for rr in (current_row-1, current_row):
                for cc in range(1, total_cols+1):
                    ws.cell(row=rr, column=cc).border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # merge 'Giorno' su tutto il blocco (ore * 2 righe)
        day_end_row = current_row
        ws.merge_cells(start_row=day_start_row, start_column=1, end_row=day_end_row, end_column=1)
        ws.cell(row=day_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

        # bordo inferiore doppio (separatore giorno) sull’ultima riga del blocco
        for c in range(1, total_cols+1):
            cell = ws.cell(row=day_end_row, column=c)
            cell.border = Border(left=cell.border.left or thin,
                                 right=cell.border.right or thin,
                                 top=cell.border.top or thin,
                                 bottom=medium)

    # ---- larghezze colonne ----
    ws.column_dimensions[get_column_letter(1)].width = float(xlsx_giorno_col_width)
    ws.column_dimensions[get_column_letter(2)].width = float(xlsx_ora_col_width)
    for c in range(3, total_cols+1):
        ws.column_dimensions[get_column_letter(c)].width = float(xlsx_aule_col_width)

    # wrap al centro per le celle delle aule
    for r in range(4, current_row+1):
        for c in range(3, total_cols+1):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # salva
    xlsx_path = OUTPUT_DIR / nome_file_xlsx
    wb.save(xlsx_path)
    print("Creato XLSX:", xlsx_path)
    return xlsx_path




# ================================
# Loader Tabella_Aule e Tabella_Classi
# ================================
def load_aule_capienze(tabella_aule_path, sheet_name=0):
    """
    Legge Tabella_Aule e restituisce un DataFrame con colonne:
    ['Aula','Plesso','Capienza'].
    Riconosce le intestazioni anche se hanno nomi leggermente diversi.
    """
    import pandas as pd
    import re
    from pathlib import Path

    # prova a trovare la riga di header "migliore" tra le prime 10
    raw = pd.read_excel(tabella_aule_path, sheet_name=sheet_name, header=None, dtype=str, engine="openpyxl")
    header_idx, best_score = 0, -1
    CAND_AULA  = {"aula","aule","nome aula","sala","room","auditorium"}
    CAND_PLESSO= {"plesso","sede","edificio","campus","building"}
    CAND_CAP   = {"capienza","capienza max","posti","posti_max","capacity","seats"}

    for i in range(min(10, len(raw))):
        row = [str(x).strip() for x in raw.iloc[i].fillna("").tolist()]
        norm = [" ".join(re.sub(r"\s+"," ",x).lower().split()) for x in row]
        score = sum(c in CAND_AULA for c in norm) + sum(c in CAND_PLESSO for c in norm) + sum(c in CAND_CAP for c in norm)
        if score > best_score:
            best_score = score
            header_idx = i

    df = pd.read_excel(tabella_aule_path, sheet_name=sheet_name, header=header_idx, engine="openpyxl")
    cols_orig = [str(c) for c in df.columns]
    cols_norm = [" ".join(re.sub(r"\s+"," ",c).lower().split()) for c in cols_orig]

    # mappa robusta delle tre colonne chiave
    col_map = {}
    for i, c in enumerate(cols_norm):
        if c in CAND_AULA   and "Aula"     not in col_map: col_map["Aula"] = cols_orig[i]
        if c in CAND_PLESSO and "Plesso"   not in col_map: col_map["Plesso"] = cols_orig[i]
        if c in CAND_CAP    and "Capienza" not in col_map: col_map["Capienza"] = cols_orig[i]

    # fallback: match "contiene" se non trovate prima
    def _fallback(target_set, key):
        if key in col_map: return
        for i, c in enumerate(cols_norm):
            if any(tok in c for tok in target_set):
                col_map[key] = cols_orig[i]; return
    _fallback(CAND_AULA,   "Aula")
    _fallback(CAND_PLESSO, "Plesso")
    _fallback(CAND_CAP,    "Capienza")

    missing = {"Aula","Plesso","Capienza"} - set(col_map.keys())
    if missing:
        from pathlib import Path
        raise ValueError(f"{Path(tabella_aule_path).name}: mancano colonne {sorted(missing)}. Intestazioni viste: {cols_orig}")

    out = df[[col_map["Aula"], col_map["Plesso"], col_map["Capienza"]]].copy()
    out.columns = ["Aula","Plesso","Capienza"]

    def tidy(x):
        if pd.isna(x): return ""
        s = str(x).strip()
        return "" if s.lower()=="nan" else s

    out["Aula"]   = out["Aula"].apply(tidy)
    out["Plesso"] = out["Plesso"].apply(tidy)

    def _to_int(x):
        s = tidy(x)
        if s == "": return None
        import re
        m = re.search(r"\d+", s)
        return int(m.group(0)) if m else None
    out["Capienza"] = out["Capienza"].apply(_to_int)

    # filtra righe senza nome aula
    out = out[out["Aula"]!=""].reset_index(drop=True)
    return out


def load_tabella_classi(tabella_classi_path, sheet_name=0):
    """
    Legge Tabella_Classi e restituisce DataFrame con colonne ['Edificio','Classe'].
    Riconosce 'Edificio/Plesso/Sede' e 'Classe' come intestazioni.
    """
    import pandas as pd
    from pathlib import Path

    raw = pd.read_excel(tabella_classi_path, sheet_name=sheet_name, header=None, dtype=str, engine="openpyxl")
    header_idx = 0
    for i in range(min(10, len(raw))):
        row = [str(x).strip().lower() for x in raw.iloc[i].fillna("").tolist()]
        if any(k in row for k in ["edificio","plesso","sede"]) and any("classe" in k for k in row):
            header_idx = i; break

    df = pd.read_excel(tabella_classi_path, sheet_name=sheet_name, header=header_idx, engine="openpyxl")
    df = df.rename(columns={c: c.strip() for c in df.columns})

    col_map = {}
    for c in df.columns:
        lc = c.strip().lower()
        if lc in {"edificio","plesso","sede"}: col_map[c] = "Edificio"
        elif lc == "classe":                   col_map[c] = "Classe"
    df = df.rename(columns=col_map)

    if not {"Edificio","Classe"} <= set(df.columns):
        raise ValueError(f"{Path(tabella_classi_path).name}: servono colonne 'Edificio' e 'Classe'. Trovate: {list(df.columns)}")

    def tidy(x):
        if pd.isna(x): return ""
        s = str(x).strip()
        return "" if s.lower()=="nan" else s

    df["Edificio"] = df["Edificio"].apply(tidy)
    df["Classe"]   = df["Classe"].apply(tidy)
    df = df[(df["Classe"]!="")].reset_index(drop=True)
    return df[["Edificio","Classe"]]


"""
AUTO-ESECUZIONE DISABILITATA PER AMBIENTE SERVER
La sezione originale eseguiva automaticamente il caricamento dati ed export.
E' stata resa inattiva per permettere l'import del modulo senza side-effect.
Usare la funzione main(input_paths, output_dir, **options) definita sotto.
"""

# ================================
# CARICAMENTO DATI (DISABILITATO: ESEMPI ORIGINALI)
# ================================
# df_aule   = load_aule_capienze(TABELLA_AULE_XLSX)
# df_centrale, giorni_c, ore_c = read_teacher_matrix(
#     CENTRALE_XLSX, plesso_label="Centrale", df_aule=df_aule
# )
# df_succ,     giorni_s, ore_s = read_teacher_matrix(
#     SUCCURSALE_XLSX, plesso_label="Succursale", df_aule=df_aule
# )
# df_classi = load_tabella_classi(TABELLA_CLASSI_XLSX)
# materie_map = load_tabella_materie(TABELLA_MATERIE_XLSX)
# if (giorni_c == giorni_s) and (ore_c == ore_s):
#     giorni, ore = giorni_c, ore_c
# else:
#     giorni = list(OrderedDict.fromkeys(list(giorni_c) + list(giorni_s)))
#     ore    = list(OrderedDict.fromkeys(list(ore_c)    + list(ore_s)))
# df_all = pd.concat([df_centrale, df_succ], ignore_index=True)
# df_sostegno, giorni_sost, ore_sost = load_tabella_sostegno(TABELLA_SOSTEGNO_XLSX, df_aule=df_aule)
# df_all = integrate_sostegno_and_mark(df_all, df_sostegno, df_aule=df_aule, df_classi=df_classi)



import time, gc, traceback

DELAY_SECONDS = 1   # pausa tra file (puoi alzare a 2–3 se serve)
RETRY_COUNT   = 1   # ritenti in caso d'errore

def run_with_delay(fn, *args, delay=DELAY_SECONDS, retries=RETRY_COUNT, **kwargs):
    """
    Esegue fn(*args, **kwargs), poi aspetta 'delay' secondi.
    In caso di eccezione, fa un retry dopo una pausa.
    Ritorna il valore di fn.
    """
    for attempt in range(retries + 1):
        try:
            result = fn(*args, **kwargs)
            gc.collect()
            time.sleep(delay)
            return result
        except Exception as e:
            print(f"[WARN] Export fallito al tentativo {attempt+1}/{retries+1}: {fn.__name__} -> {e}")
            traceback.print_exc()
            if attempt < retries:
                time.sleep(max(delay, 2.0))
            else:
                raise


# ================================
# COSTRUZIONE E ESPORTAZIONE (DISABILITATO A LIVELLO MODULO)
# Eseguire solo dentro main()
# ================================
# xlsx1 = run_with_delay(
#     export_OUTPUT_CLASSI_SETTIMANALE, df_all, df_classi,
#     titolo="ORARIO_CLASSI_SETTIMANALE", materie_map=materie_map
# )

# xlsx_as_c = run_with_delay(
#     export_OUTPUT_AULE_SETTIMANALE, df_all, df_aule,
#     titolo="ORARIO_AULE_SETTIMANALE",
#     plesso=None,
#     xlsx_first_col_width=6.5,
#     xlsx_second_col_width=10.0,
#     xlsx_day_col_width=None
# )

# xlsx2 = run_with_delay(
#     export_OUTPUT_TABELLA_GLOBALE, df_all, giorni, ore,
#     df_aule=df_aule, df_classi=df_classi
# )

## ESEMPIO DISABILITATO
# xlsx_aule_s = run_with_delay(
#     export_OUTPUT_AULE_COMPATTO, df_succ, "ORARIO_AULE_COMPATTO_SUCCURSALE.xlsx",
#     df_aule=df_aule,
#     # xlsx_first_col_width=18.0, xlsx_time_col_width=18.0
# )


## ESEMPIO DISABILITATO
# xlsx_aule_c = run_with_delay(
#     export_OUTPUT_AULE_COMPATTO, df_centrale, "ORARIO_AULE_COMPATTO_CENTRALE.xlsx",
#     df_aule=df_aule,
#     # xlsx_first_col_width=18.0, xlsx_time_col_width=18.0
# )


## ESEMPIO DISABILITATO
# xlsx5 = run_with_delay(
#     export_OUTPUT_CLASSI_COMPATTO, df_succ, "ORARIO_CLASSI_COMPATTO_SUCCURSALE.xlsx",
#     # xlsx_first_col_width=16.0, xlsx_time_col_width=16.0
# )

## ESEMPIO DISABILITATO
# xlsx6 = run_with_delay(
#     export_OUTPUT_CLASSI_COMPATTO, df_centrale, "ORARIO_CLASSI_COMPATTO_CENTRALE.xlsx",
#     # xlsx_first_col_width=16.0, xlsx_time_col_width=16.0
# )

## ESEMPIO DISABILITATO
# xlsx_loc_s = run_with_delay(
#     export_OUTPUT_TABELLA_PLESSO, df_all, giorni, ore,
#     df_aule=df_aule, plesso_focus="Succursale",
#     # docenti_set=docenti_set,
#     df_classi=df_classi
# )

## ESEMPIO DISABILITATO
# xlsx_loc_c = run_with_delay(
#     export_OUTPUT_TABELLA_PLESSO, df_all, giorni, ore,
#     df_aule=df_aule, plesso_focus="Centrale",
#     # docenti_set=docenti_set,
#     df_classi=df_classi
# )

## ESEMPIO DISABILITATO
# export_OUTPUT_AULE_LIBERE(
#     df_centrale, df_aule, "Centrale", "ORARIO_AULE_LIBERE_CENTRALE.xlsx",
#     xlsx_giorno_col_width=11, xlsx_ora_col_width=6,
# )

## ESEMPIO DISABILITATO
# export_OUTPUT_AULE_LIBERE(
#     df_succ, df_aule, "Succursale", "ORARIO_AULE_LIBERE_SUCCURSALE.xlsx",
#     xlsx_giorno_col_width=11, xlsx_ora_col_width=6,
# )

# print("\nTutti gli XLSX sono in:", OUTPUT_DIR)

# import shutil
# from google.colab import files

# # Crea un archivio ZIP con tutti i file prodotti
# archive_base = "/content/DADA_outputs"
# shutil.make_archive(archive_base, "zip", str(OUTPUT_DIR))

# # Scarica lo zip sul tuo computer
# # files.download(archive_base + ".zip")


# ================================
# Entrypoint per integrazione con FastAPI adapter
# ================================
def main(input_paths: list[str], output_dir: str, **options) -> int:
    """
    Entry point compatibile con l'app.
    - input_paths: lista file caricati via UI (.xlsx)
    - output_dir: cartella di lavoro del job dove scrivere TUTTI gli output
    - options: parametri opzionali (es. header_text)
    """
    from pathlib import Path
    import traceback
    try:
        # 1) Setup output dir e header
        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)
        global OUTPUT_DIR
        OUTPUT_DIR = out

        # HEADER_TEXT dalle options (UI)
        hdr = (options or {}).get("header_text") or (options or {}).get("HEADER_TEXT") or ""
        try:
            set_header_text(hdr)
        except Exception:
            pass

        # 2) Mappa input per nome atteso (case-insensitive)
        name_map = {Path(p).name.lower(): Path(p) for p in input_paths}
        def pick(*candidates: str) -> Path | None:
            for c in candidates:
                p = name_map.get(c.lower())
                if p: return p
            return None

        centrale   = pick("centrale.xlsx")
        succursale = pick("succursale.xlsx")
        tab_aule   = pick("tabella_aule.xlsx", "aule.xlsx")
        tab_classi = pick("tabella_classi.xlsx", "classi.xlsx")
        tab_materie= pick("tabella_materie.xlsx", "materie.xlsx")
        tab_sost   = pick("tabella_sostegno.xlsx", "sostegno.xlsx")

        missing = [n for n,p in {
            "Centrale.xlsx": centrale,
            "Succursale.xlsx": succursale,
            "Tabella_Aule.xlsx": tab_aule,
            "Tabella_Classi.xlsx": tab_classi,
            "Tabella_Materie.xlsx": tab_materie,
            "Tabella_Sostegno.xlsx": tab_sost,
        }.items() if p is None]
        if missing:
            (out / "ERROR_MISSING_INPUTS.txt").write_text(
                "Mancano i seguenti file richiesti:\n- " + "\n- ".join(missing) +
                "\nCarica tutti i file necessari dalla UI.", encoding="utf-8"
            )
            return 1

        # 3) Caricamento e pipeline (equivalente sezione originale)
        df_aule = load_aule_capienze(tab_aule)
        df_centrale, giorni_c, ore_c = read_teacher_matrix(centrale, plesso_label="Centrale", df_aule=df_aule)
        df_succ,     giorni_s, ore_s = read_teacher_matrix(succursale, plesso_label="Succursale", df_aule=df_aule)

        df_classi = load_tabella_classi(tab_classi)
        materie_map = load_tabella_materie(tab_materie)

        if (giorni_c == giorni_s) and (ore_c == ore_s):
            giorni, ore = giorni_c, ore_c
        else:
            giorni = list(OrderedDict.fromkeys(list(giorni_c) + list(giorni_s)))
            ore    = list(OrderedDict.fromkeys(list(ore_c)    + list(ore_s)))

        df_all = pd.concat([df_centrale, df_succ], ignore_index=True)
        df_sostegno, _, _ = load_tabella_sostegno(tab_sost, df_aule=df_aule)
        df_all = integrate_sostegno_and_mark(df_all, df_sostegno, df_aule=df_aule, df_classi=df_classi)

        # 4) Export principali (solo XLSX)
        run_with_delay(export_OUTPUT_CLASSI_SETTIMANALE, df_all, df_classi, titolo="ORARIO_CLASSI_SETTIMANALE", materie_map=materie_map)
        run_with_delay(export_OUTPUT_AULE_SETTIMANALE,   df_all, df_aule, titolo="ORARIO_AULE_SETTIMANALE", plesso=None,
                       xlsx_first_col_width=6.5, xlsx_second_col_width=10.0, xlsx_day_col_width=None)
        run_with_delay(export_OUTPUT_TABELLA_GLOBALE,   df_all, giorni, ore, df_aule=df_aule, df_classi=df_classi)
        run_with_delay(export_OUTPUT_AULE_COMPATTO,     df_succ,     "ORARIO_AULE_COMPATTO_SUCCURSALE.xlsx", df_aule=df_aule)
        run_with_delay(export_OUTPUT_AULE_COMPATTO,     df_centrale, "ORARIO_AULE_COMPATTO_CENTRALE.xlsx",  df_aule=df_aule)
        run_with_delay(export_OUTPUT_CLASSI_COMPATTO,   df_succ,     "ORARIO_CLASSI_COMPATTO_SUCCURSALE.xlsx")
        run_with_delay(export_OUTPUT_CLASSI_COMPATTO,   df_centrale, "ORARIO_CLASSI_COMPATTO_CENTRALE.xlsx")
        run_with_delay(export_OUTPUT_TABELLA_PLESSO,    df_all, giorni, ore, df_aule=df_aule, plesso_focus="Succursale", df_classi=df_classi)
        run_with_delay(export_OUTPUT_TABELLA_PLESSO,    df_all, giorni, ore, df_aule=df_aule, plesso_focus="Centrale",  df_classi=df_classi)
        export_OUTPUT_AULE_LIBERE(df_centrale, df_aule, "Centrale",   "ORARIO_AULE_LIBERE_CENTRALE.xlsx",  xlsx_giorno_col_width=11, xlsx_ora_col_width=6)
        export_OUTPUT_AULE_LIBERE(df_succ,     df_aule, "Succursale", "ORARIO_AULE_LIBERE_SUCCURSALE.xlsx", xlsx_giorno_col_width=11, xlsx_ora_col_width=6)

        # 5) Report finale
        (out / "_OK.txt").write_text("Export completato.", encoding="utf-8")
        return 0
    except Exception as e:
        (Path(output_dir) / "_ERROR.txt").write_text(str(e) + "\n" + traceback.format_exc(), encoding="utf-8")
        return 1

