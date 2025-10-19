"""
Optional schema validation for Excel files.
Can be extended to check required sheets, headers, data types, etc.
"""
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class SchemaValidator:
    """
    Validates Excel files against a schema.
    Example schema:
    {
        "required_sheets": ["Classi", "Studenti"],
        "sheets": {
            "Classi": {"required_headers": ["ID", "Nome"]},
            "Studenti": {"required_headers": ["ID", "Nome", "Classe"]}
        }
    }
    """

    def __init__(self, schema: Optional[dict] = None):
        self.schema = schema or {}

    def validate(self, file_path: Path) -> tuple[bool, str]:
        """
        Validate a file against the schema.
        Returns (is_valid, error_message).
        """
        if not HAS_OPENPYXL:
            return True, ""  # Skip if openpyxl not installed

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            return False, f"Impossibile aprire file: {e}"

        # Check required sheets
        required_sheets = self.schema.get("required_sheets", [])
        sheet_names = set(wb.sheetnames)
        missing = set(required_sheets) - sheet_names
        if missing:
            return False, f"Fogli mancanti: {', '.join(missing)}"

        # Check sheet headers
        sheets_config = self.schema.get("sheets", {})
        for sheet_name, config in sheets_config.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            required_headers = config.get("required_headers", [])
            if required_headers:
                first_row = [cell.value for cell in ws[1]]
                missing_headers = set(required_headers) - set(first_row)
                if missing_headers:
                    return False, f"Foglio '{sheet_name}': intestazioni mancanti: {', '.join(missing_headers)}"

        return True, ""


# Default schema (can be overridden via env or config)
DEFAULT_SCHEMA = {
    "required_sheets": [],  # No required sheets by default
    "sheets": {}
}

validator = SchemaValidator(DEFAULT_SCHEMA)
